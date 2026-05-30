"""
App KBC-HP89: hệ thống quản lý liên kết KBC <-> HP89
- Đặt hàng (workflow 5 bước HP89 -> KBC)
- Hợp đồng / Thỏa thuận giữa KBC và HP89
- Giấy tờ Pháp lý
- Truyền thông (HP89 đăng nội dung -> KBC duyệt)
- Báo cáo Excel
- Quản lý người dùng (phân biệt nhân sự KBC / HP89)
Flask + SQLite + Flask-Login
"""
import os
import sqlite3
import smtplib
import io
import json
import zipfile
import tempfile
from datetime import datetime, date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for, flash,
    send_from_directory, send_file, abort, jsonify, g
)
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from dotenv import load_dotenv

load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get('DATA_DIR', BASE_DIR)
UPLOAD_DIR = os.path.join(DATA_DIR, 'uploads')
ORDER_DIR = os.path.join(UPLOAD_DIR, 'orders')
CONTRACT_DIR = os.path.join(UPLOAD_DIR, 'contracts')
ATTACH_DIR = os.path.join(UPLOAD_DIR, 'attachments')
LEGAL_DIR = os.path.join(UPLOAD_DIR, 'legal')
MEDIA_DIR = os.path.join(UPLOAD_DIR, 'media')
DB_PATH = os.path.join(DATA_DIR, 'data.db')

for d in (DATA_DIR, UPLOAD_DIR, ORDER_DIR, CONTRACT_DIR, ATTACH_DIR, LEGAL_DIR, MEDIA_DIR):
    os.makedirs(d, exist_ok=True)

print(f'>>> DATABASE PATH: {DB_PATH}', flush=True)
print(f'>>> DATA_DIR env: {os.environ.get("DATA_DIR", "(not set — using BASE_DIR)")}', flush=True)

ALLOWED_EXT = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'png', 'jpg', 'jpeg', 'zip', 'rar', 'txt', 'mp4', 'mp3'}
MAX_UPLOAD_MB = 100

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-me-in-production-kbc-hp89')
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024

VAPID_PUBLIC_KEY = os.environ.get('VAPID_PUBLIC_KEY', '')
VAPID_PRIVATE_KEY = os.environ.get('VAPID_PRIVATE_KEY', '')
VAPID_CLAIM_EMAIL = os.environ.get('VAPID_CLAIM_EMAIL', 'mailto:admin@example.com')

# URL tuyệt đối của app — dùng để chèn vào email (Render set RENDER_EXTERNAL_URL tự động).
APP_BASE_URL = (os.environ.get('APP_BASE_URL')
                or os.environ.get('RENDER_EXTERNAL_URL', '')).rstrip('/')

# SMTP gửi email — bỏ qua êm nếu chưa cấu hình
SMTP_HOST = os.environ.get('SMTP_HOST')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER')
SMTP_PASS = os.environ.get('SMTP_PASS')
SMTP_FROM = os.environ.get('SMTP_FROM', SMTP_USER or 'noreply@kbc-hp89.vn')
SMTP_FROM_NAME = os.environ.get('SMTP_FROM_NAME', 'KBC-HP89')

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Vui lòng đăng nhập để tiếp tục.'


# ---------- DB ----------
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        full_name TEXT,
        email TEXT,
        role TEXT NOT NULL DEFAULT 'staff',
        organization TEXT NOT NULL DEFAULT 'HP89',  -- 'KBC' | 'HP89'
        permissions TEXT,
        created_at TEXT NOT NULL
    );

    -- Thông tin đơn vị xuất hóa đơn (lưu sẵn để tái sử dụng)
    CREATE TABLE IF NOT EXISTS invoice_entities (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_name TEXT NOT NULL,
        tax_code TEXT,
        address TEXT,
        email TEXT,
        phone TEXT,
        created_by INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(created_by) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_invoice_entities_creator ON invoice_entities(created_by);

    -- Danh mục sản phẩm
    CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        unit TEXT,
        packaging TEXT,
        default_price REAL DEFAULT 0,
        active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL
    );

    -- Đơn đặt hàng (HP89 đặt -> KBC giao)
    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT,
        order_date TEXT,
        customer_name TEXT,        -- Nơi nhận hàng (cửa hàng/chi nhánh HP89)
        customer_phone TEXT,
        customer_address TEXT,
        invoice_company TEXT,
        invoice_tax_code TEXT,
        invoice_address TEXT,
        invoice_email TEXT,
        invoice_phone TEXT,
        payment_method TEXT,
        bank_account TEXT,
        signer_creator TEXT,
        signer_approver TEXT,
        signer_deliverer TEXT,
        subtotal REAL DEFAULT 0,
        vat_percent REAL DEFAULT 8,
        vat_amount REAL DEFAULT 0,
        grand_total REAL DEFAULT 0,
        notes TEXT,
        order_file TEXT,
        invoice_file TEXT,
        delivery_file TEXT,
        other_file TEXT,
        warehouse_file TEXT,
        paid_amount REAL DEFAULT 0,
        referrer TEXT,
        -- Workflow 5 trạng thái:
        -- draft: HP89 lưu nháp
        -- pending_hp89: HP89 đã gửi, chờ lãnh đạo HP89 duyệt
        -- approved_hp89: Lãnh đạo HP89 đã duyệt, đẩy sang KBC
        -- received_kbc: Nhân viên KBC đã nhận đơn
        -- delivered_kbc: Nhân viên KBC đã giao hàng (HOÀN THÀNH)
        workflow_status TEXT DEFAULT 'draft',
        submitted_at TEXT,
        approved_by INTEGER,
        approved_at TEXT,
        received_by INTEGER,
        received_at TEXT,
        delivered_by INTEGER,
        delivered_at TEXT,
        owner_id INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS order_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        product_id INTEGER,
        product_name TEXT NOT NULL,
        unit TEXT,
        packaging TEXT,
        qty REAL DEFAULT 0,
        unit_price REAL DEFAULT 0,
        amount REAL DEFAULT 0,
        discount REAL DEFAULT 0,
        amount_after REAL DEFAULT 0,
        note TEXT,
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY(order_id) REFERENCES orders(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_order_items_order ON order_items(order_id);
    CREATE INDEX IF NOT EXISTS idx_orders_date ON orders(order_date);
    CREATE INDEX IF NOT EXISTS idx_orders_workflow ON orders(workflow_status);

    -- Hợp đồng / Thỏa thuận giữa KBC & HP89
    CREATE TABLE IF NOT EXISTS contracts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT,
        title TEXT NOT NULL,
        partner TEXT,             -- Đơn vị đối tác (mặc định KBC hoặc HP89)
        supplier_tax TEXT,
        contact_person TEXT,
        contact_phone TEXT,
        receiver TEXT,
        receiver_email TEXT,
        receiver_ids TEXT,
        total_value REAL DEFAULT 0,
        paid_amount REAL DEFAULT 0,
        contract_date TEXT,
        due_date TEXT,
        progress TEXT,
        progress_percent INTEGER DEFAULT 0,
        notes TEXT,
        contract_file TEXT,
        handover_file TEXT,
        appendix_file TEXT,
        invoice_file TEXT,
        other_file TEXT,
        owner_id INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_contract_due ON contracts(due_date);

    -- Nhận xét cho contract/order/media
    CREATE TABLE IF NOT EXISTS comments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_type TEXT NOT NULL,
        record_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        content TEXT NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_comments_record ON comments(record_type, record_id);

    CREATE TABLE IF NOT EXISTS record_permissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_type TEXT NOT NULL,
        record_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        perm TEXT NOT NULL,
        UNIQUE(record_type, record_id, user_id),
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_perm_lookup ON record_permissions(record_type, record_id, user_id);

    -- File đính kèm cho contract
    CREATE TABLE IF NOT EXISTS attachments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_type TEXT NOT NULL,
        record_id INTEGER NOT NULL,
        category TEXT NOT NULL,
        stored_name TEXT NOT NULL,
        original_name TEXT,
        created_at TEXT NOT NULL
    );
    CREATE INDEX IF NOT EXISTS idx_attach_lookup ON attachments(record_type, record_id, category);

    -- Cây giấy tờ pháp lý
    CREATE TABLE IF NOT EXISTS legal_nodes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        parent_id INTEGER,
        description TEXT,
        stored_name TEXT,
        original_name TEXT,
        approval_status TEXT DEFAULT 'pending',
        approval_note TEXT,
        approved_by INTEGER,
        approved_at TEXT,
        created_by INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(parent_id) REFERENCES legal_nodes(id) ON DELETE CASCADE,
        FOREIGN KEY(created_by) REFERENCES users(id) ON DELETE CASCADE,
        FOREIGN KEY(approved_by) REFERENCES users(id) ON DELETE SET NULL
    );
    CREATE INDEX IF NOT EXISTS idx_legal_parent ON legal_nodes(parent_id);

    -- Truyền thông: HP89 đăng nội dung, KBC duyệt
    CREATE TABLE IF NOT EXISTS media_posts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        content TEXT,
        channel TEXT,             -- Báo chí / Facebook / Website / TikTok / ...
        planned_date TEXT,        -- Ngày dự kiến đăng
        link TEXT,                -- Link bài viết (nếu đã đăng)
        notes TEXT,
        -- Trạng thái:
        -- draft: HP89 đang soạn
        -- pending: HP89 đã gửi, chờ KBC duyệt
        -- approved: KBC đã đồng ý thống nhất nội dung
        -- revision: KBC yêu cầu sửa
        -- published: Đã đăng
        status TEXT DEFAULT 'draft',
        revision_note TEXT,
        confirmed_by INTEGER,
        confirmed_at TEXT,
        submitted_at TEXT,
        owner_id INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS media_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        post_id INTEGER NOT NULL,
        stored_name TEXT NOT NULL,
        original_name TEXT,
        uploaded_by INTEGER,
        created_at TEXT NOT NULL,
        FOREIGN KEY(post_id) REFERENCES media_posts(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_media_files_post ON media_files(post_id);
    CREATE INDEX IF NOT EXISTS idx_media_status ON media_posts(status);

    CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        message TEXT NOT NULL,
        link TEXT,
        is_read INTEGER DEFAULT 0,
        created_at TEXT NOT NULL,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_notif_user ON notifications(user_id, is_read);

    CREATE TABLE IF NOT EXISTS push_subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        endpoint TEXT UNIQUE NOT NULL,
        subscription TEXT NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_push_user ON push_subscriptions(user_id);
    ''')

    # Tạo admin mặc định nếu chưa có
    cur = conn.execute('SELECT COUNT(*) FROM users WHERE role=?', ('admin',))
    if cur.fetchone()[0] == 0:
        conn.execute(
            'INSERT INTO users(username, password_hash, full_name, role, organization, created_at) VALUES (?,?,?,?,?,?)',
            ('admin', generate_password_hash('admin123'), 'Quản trị viên', 'admin', 'KBC',
             datetime.now().isoformat(timespec='seconds'))
        )
        print('>>> Default admin created: admin / admin123 (CHANGE PASSWORD AFTER FIRST LOGIN!)')
    conn.commit()
    conn.close()


# ---------- Vai trò & Tổ chức ----------
ROLE_LABELS = {
    'admin': 'Admin',
    'director': 'Giám đốc',
    'deputy_director': 'Phó Giám đốc',
    'manager': 'Trưởng phòng / Kiểm soát',
    'staff': 'Nhân viên',
    'board': 'Thành viên HĐQT',
}
ROLE_ORDER = ['admin', 'director', 'deputy_director', 'manager', 'staff', 'board']

ORGS = ['KBC', 'HP89']
ORG_LABELS = {
    'KBC': 'KBC',
    'HP89': 'HP89',
}

FULL_ACCESS_ROLES = ('admin',)

# Capabilities cho phân quyền tích chọn
CAPABILITIES = [
    ('view_all', 'Xem TẤT CẢ đơn hàng, hợp đồng, truyền thông của mọi người'),
    ('create_order', 'Tạo đơn đặt hàng (HP89)'),
    ('approve_hp89_order', 'Duyệt đơn đặt hàng phía HP89 (Lãnh đạo HP89)'),
    ('receive_kbc_order', 'Nhận đơn hàng phía KBC (Nhân viên KBC)'),
    ('deliver_kbc_order', 'Tick đã giao hàng phía KBC (Nhân viên KBC)'),
    ('manage_contract', 'Tạo/sửa hợp đồng & thỏa thuận'),
    ('manage_products', 'Quản lý danh mục sản phẩm'),
    ('create_media', 'Đăng nội dung truyền thông (HP89)'),
    ('approve_media', 'Duyệt/Xác nhận nội dung truyền thông (KBC)'),
    ('approve_legal', 'Phê duyệt Giấy tờ Pháp lý'),
    ('comment', 'Nhận xét trên đơn hàng / hợp đồng / truyền thông'),
    ('notify_order', 'Nhận thông báo khi có đơn hàng mới hoặc thay đổi trạng thái'),
    ('notify_contract', 'Nhận thông báo khi có hợp đồng mới'),
    ('notify_media', 'Nhận thông báo khi có bài truyền thông mới'),
]

MANAGE_USERS_CAP = ('manage_users', 'Quản lý người dùng & phân quyền (tạo/sửa/xoá user)')
CAP_KEYS = {c[0] for c in CAPABILITIES} | {MANAGE_USERS_CAP[0]}
ALL_NOTIFY_CAPS = {'notify_order', 'notify_contract', 'notify_media'}

# Cap nào thuộc tổ chức nào (None = dùng chung cho cả 2)
# Chỉ admin có thể cấp/dùng mọi cap. Manager bên KBC/HP89 chỉ cấp được cap
# của tổ chức mình + các cap chung.
CAP_ORGS = {
    # KBC-only
    'receive_kbc_order': 'KBC',
    'deliver_kbc_order': 'KBC',
    'approve_media': 'KBC',
    'manage_products': 'KBC',
    # HP89-only
    'create_order': 'HP89',
    'approve_hp89_order': 'HP89',
    'create_media': 'HP89',
    # Dùng chung
    'view_all': None,
    'manage_contract': None,
    'approve_legal': None,
    'comment': None,
    'notify_order': None,
    'notify_contract': None,
    'notify_media': None,
    'manage_users': None,
}


def cap_for_org(cap, org):
    """Cap có phù hợp với tổ chức org không?
    None trong CAP_ORGS = cap chung, ai cũng cấp được."""
    cap_org = CAP_ORGS.get(cap)
    return cap_org is None or cap_org == org


def caps_for_org(org):
    """Tập tất cả cap mà user thuộc org có thể được cấp (kể cả cap chung)."""
    return {c for c in CAP_KEYS if cap_for_org(c, org)}

# Quyền mặc định theo vai trò
DEFAULT_CAPS = {
    'director': {'view_all', 'approve_hp89_order', 'manage_contract', 'manage_products',
                 'approve_media', 'approve_legal', 'comment'} | ALL_NOTIFY_CAPS,
    'deputy_director': {'view_all', 'approve_hp89_order', 'manage_contract',
                        'approve_media', 'approve_legal', 'comment'} | ALL_NOTIFY_CAPS,
    'manager': {'view_all', 'create_order', 'create_media', 'manage_contract', 'comment'} | ALL_NOTIFY_CAPS,
    'staff': {'create_order', 'receive_kbc_order', 'deliver_kbc_order', 'create_media', 'comment'},
    'board': {'view_all'} | ALL_NOTIFY_CAPS,
}


# ---------- User ----------
class User(UserMixin):
    def __init__(self, row):
        self.id = row['id']
        self.username = row['username']
        self.full_name = row['full_name']
        self.email = row['email']
        self.role = row['role']
        keys = row.keys()
        self.organization = row['organization'] if 'organization' in keys else 'HP89'
        self.permissions = row['permissions'] if 'permissions' in keys else None

    @property
    def is_admin(self):
        return self.role == 'admin'

    @property
    def full_access(self):
        return self.role in FULL_ACCESS_ROLES

    @property
    def is_kbc(self):
        return self.organization == 'KBC'

    @property
    def is_hp89(self):
        return self.organization == 'HP89'

    @property
    def can_manage_users(self):
        return self.has_cap('manage_users')

    @property
    def _granted(self):
        if not self.permissions:
            return set()
        return {c.strip() for c in self.permissions.split(',') if c.strip()}

    @property
    def caps(self):
        """Tập quyền hiệu lực:
        - Admin: tất cả CAP_KEYS.
        - User khác: caps được cấp (hoặc default theo role) — đã LỌC bỏ những cap
          không thuộc tổ chức của user (KBC user không bao giờ có cap HP89-only và
          ngược lại, bất kể DB có gì)."""
        if self.is_admin:
            return set(CAP_KEYS)
        if self.permissions is None:
            raw = set(DEFAULT_CAPS.get(self.role, set()))
        else:
            raw = self._granted
        # Lọc theo tổ chức: chỉ giữ cap chung hoặc cap thuộc đúng org
        return {c for c in raw if CAP_ORGS.get(c) in (None, self.organization)}

    def has_cap(self, cap):
        return cap in self.caps

    @property
    def can_see_all(self):
        return self.full_access or self.has_cap('view_all')

    @property
    def role_label(self):
        return ROLE_LABELS.get(self.role, self.role)

    @property
    def org_label(self):
        return ORG_LABELS.get(self.organization, self.organization)


@login_manager.user_loader
def load_user(user_id):
    row = get_db().execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
    return User(row) if row else None


def admin_required(f):
    @wraps(f)
    def wrap(*a, **kw):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return f(*a, **kw)
    return wrap


def manage_users_required(f):
    @wraps(f)
    def wrap(*a, **kw):
        if not current_user.is_authenticated or not current_user.can_manage_users:
            abort(403)
        return f(*a, **kw)
    return wrap


def require_cap(cap):
    def deco(f):
        @wraps(f)
        def wrap(*a, **kw):
            if not current_user.is_authenticated or not current_user.has_cap(cap):
                abort(403)
            return f(*a, **kw)
        return wrap
    return deco


# ---------- Permission helpers ----------
def get_perm(record_type, record_id, user):
    if user.full_access:
        return 'owner'
    table = {'contract': 'contracts', 'order': 'orders', 'media': 'media_posts'}.get(record_type)
    if not table:
        return None
    row = get_db().execute(f'SELECT owner_id FROM {table} WHERE id=?', (record_id,)).fetchone()
    if not row:
        return None
    if row['owner_id'] == user.id:
        return 'owner'
    perm = get_db().execute(
        'SELECT perm FROM record_permissions WHERE record_type=? AND record_id=? AND user_id=?',
        (record_type, record_id, user.id)
    ).fetchone()
    if perm:
        return perm['perm']
    if user.can_see_all:
        return 'view'
    return None


def can_view(record_type, record_id, user):
    return get_perm(record_type, record_id, user) in ('owner', 'edit', 'view')


def can_edit(record_type, record_id, user):
    return get_perm(record_type, record_id, user) in ('owner', 'edit')


def list_accessible_ids(record_type, user):
    if user.can_see_all:
        return None
    table = {'contract': 'contracts', 'order': 'orders', 'media': 'media_posts'}.get(record_type)
    if not table:
        return set()
    owned = {r['id'] for r in get_db().execute(f'SELECT id FROM {table} WHERE owner_id=?', (user.id,))}
    shared = {r['record_id'] for r in get_db().execute(
        'SELECT record_id FROM record_permissions WHERE record_type=? AND user_id=?',
        (record_type, user.id))}
    return owned | shared


# ---------- Utils ----------
def allowed_file(name):
    return '.' in name and name.rsplit('.', 1)[1].lower() in ALLOWED_EXT


def save_upload(file_storage, subdir):
    if not file_storage or not file_storage.filename:
        return None, None
    if not allowed_file(file_storage.filename):
        flash(f'File không hợp lệ: {file_storage.filename}', 'danger')
        return None, None
    safe = secure_filename(file_storage.filename)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
    final_name = f'{stamp}_{safe}'
    path = os.path.join(subdir, final_name)
    file_storage.save(path)
    return final_name, file_storage.filename


def _parse_money(s):
    if not s:
        return 0.0
    try:
        return float(str(s).replace(',', '').strip())
    except ValueError:
        return 0.0


def parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None


@app.template_filter('money2')
def money2(v):
    try:
        v = float(v or 0)
    except (ValueError, TypeError):
        return v
    s = f'{v:,.2f}'
    return s.replace(',', '#').replace('.', ',').replace('#', '.')


@app.template_filter('money0')
def money0(v):
    try:
        v = float(v or 0)
    except (ValueError, TypeError):
        return v
    s = f'{v:,.0f}'
    return s.replace(',', '.')


@app.template_filter('vndate')
def vndate(s):
    if not s:
        return ''
    try:
        return datetime.strptime(s, '%Y-%m-%d').strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        return s


@app.template_filter('vndatetime')
def vndatetime(s):
    if not s:
        return ''
    try:
        return datetime.strptime(s[:19], '%Y-%m-%dT%H:%M:%S').strftime('%d/%m/%Y %H:%M')
    except (ValueError, TypeError):
        return s


def contract_status(due_date_str):
    d = parse_date(due_date_str)
    if not d:
        return 'none', None
    today = date.today()
    days = (d - today).days
    if days < 0:
        return 'overdue', days
    if days <= 7:
        return 'soon', days
    return 'ok', days


app.jinja_env.globals['contract_status'] = contract_status


# ---------- Workflow labels ----------
ORDER_WORKFLOW = ['draft', 'pending_hp89', 'approved_hp89', 'received_kbc', 'delivered_kbc']
ORDER_WF_LABEL = {
    'draft': 'Nháp HP89',
    'pending_hp89': 'Chờ Lãnh đạo HP89 duyệt',
    'approved_hp89': 'HP89 đã duyệt — Chờ KBC nhận',
    'received_kbc': 'KBC đã nhận — Chờ giao hàng',
    'delivered_kbc': 'KBC đã giao hàng — Hoàn thành',
}
ORDER_WF_PILL = {
    'draft': 'pill-gray',
    'pending_hp89': 'pill-yellow',
    'approved_hp89': 'pill-blue',
    'received_kbc': 'pill-purple',
    'delivered_kbc': 'pill-green',
}
app.jinja_env.globals['ORDER_WF_LABEL'] = ORDER_WF_LABEL
app.jinja_env.globals['ORDER_WF_PILL'] = ORDER_WF_PILL
app.jinja_env.globals['ORDER_WORKFLOW'] = ORDER_WORKFLOW

MEDIA_STATUS = ['draft', 'pending', 'approved', 'revision', 'published']
MEDIA_STATUS_LABEL = {
    'draft': 'Nháp',
    'pending': 'Chờ KBC duyệt',
    'approved': 'KBC đã thống nhất',
    'revision': 'KBC yêu cầu sửa',
    'published': 'Đã đăng',
}
MEDIA_STATUS_PILL = {
    'draft': 'pill-gray',
    'pending': 'pill-yellow',
    'approved': 'pill-green',
    'revision': 'pill-red',
    'published': 'pill-blue',
}
app.jinja_env.globals['MEDIA_STATUS_LABEL'] = MEDIA_STATUS_LABEL
app.jinja_env.globals['MEDIA_STATUS_PILL'] = MEDIA_STATUS_PILL

MEDIA_CHANNELS = ['Báo chí', 'Facebook', 'Website', 'TikTok', 'YouTube', 'Zalo', 'Khác']
app.jinja_env.globals['MEDIA_CHANNELS'] = MEDIA_CHANNELS

LEGAL_STATUS_LABEL = {
    'pending': 'Chờ duyệt',
    'approved': 'Đã duyệt',
    'rejected': 'Từ chối',
}
LEGAL_STATUS_PILL = {
    'pending': 'pill-yellow',
    'approved': 'pill-green',
    'rejected': 'pill-red',
}
app.jinja_env.globals['LEGAL_STATUS_LABEL'] = LEGAL_STATUS_LABEL
app.jinja_env.globals['LEGAL_STATUS_PILL'] = LEGAL_STATUS_PILL

app.jinja_env.globals['ORG_LABELS'] = ORG_LABELS


# ---------- Notifications ----------
def send_web_push_to_user(user_id, title, body, url=None):
    if not (user_id and VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY):
        return
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        return
    db = get_db()
    rows = db.execute('SELECT id, subscription FROM push_subscriptions WHERE user_id=?',
                      (user_id,)).fetchall()
    if not rows:
        return
    payload = json.dumps({'title': title, 'body': body, 'url': url or '/'})
    dead = []
    for r in rows:
        try:
            webpush(
                subscription_info=json.loads(r['subscription']),
                data=payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={'sub': VAPID_CLAIM_EMAIL},
            )
        except WebPushException as e:
            status = getattr(getattr(e, 'response', None), 'status_code', None)
            if status in (404, 410):
                dead.append(r['id'])
        except Exception:
            pass
    for sid in dead:
        db.execute('DELETE FROM push_subscriptions WHERE id=?', (sid,))


def create_notification(user_id, message, link=None):
    if not user_id:
        return
    db = get_db()
    db.execute('INSERT INTO notifications(user_id, message, link, is_read, created_at) VALUES (?,?,?,0,?)',
               (user_id, message, link, datetime.now().isoformat(timespec='seconds')))
    send_web_push_to_user(user_id, 'KBC-HP89', message, link)


def notify_cap_users(cap, message, link=None, exclude_ids=None, organization=None):
    """Gửi thông báo cho user có cap. Nếu organization được truyền, chỉ user thuộc org đó nhận."""
    exclude = set(exclude_ids or [])
    db = get_db()
    rows = db.execute('SELECT * FROM users').fetchall()
    for r in rows:
        if r['id'] in exclude:
            continue
        u = User(r)
        if organization and u.organization != organization and not u.is_admin:
            continue
        if u.has_cap(cap):
            create_notification(r['id'], message, link)


def notify_user(user_id, message, link=None):
    create_notification(user_id, message, link)


# ---------- Email helpers ----------
def absolute_link(relative):
    """Trả link tuyệt đối cho email (cần để bấm được từ inbox)."""
    if not relative:
        return ''
    if relative.startswith('http'):
        return relative
    if APP_BASE_URL:
        return APP_BASE_URL + relative
    return relative


def send_simple_email(to_emails, subject, body):
    """Gửi email text đơn giản. Trả (ok, msg). Bỏ qua nếu chưa cấu hình SMTP."""
    recipients = [e for e in (to_emails or []) if e and '@' in (e or '')]
    if not (SMTP_HOST and SMTP_USER and SMTP_PASS):
        return False, 'SMTP chưa cấu hình'
    if not recipients:
        return False, 'Không có email người nhận'
    msg = MIMEMultipart()
    msg['From'] = f'{SMTP_FROM_NAME} <{SMTP_FROM}>'
    msg['To'] = ', '.join(recipients)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(SMTP_FROM, recipients, msg.as_string())
    except Exception as e:
        print(f'[EMAIL ERROR] {e}')
        return False, f'Lỗi gửi email: {e}'
    return True, 'OK'


def send_email_to_user(user_id, subject, body):
    """Tra cứu email của user rồi gửi. Bỏ qua nếu user không có email."""
    if not user_id:
        return False, 'Không có user_id'
    row = get_db().execute('SELECT email, full_name, username FROM users WHERE id=?', (user_id,)).fetchone()
    if not row:
        return False, 'User không tồn tại'
    if not row['email']:
        return False, f'User "{row["full_name"] or row["username"]}" chưa có email'
    return send_simple_email([row['email']], subject, body)


def notify_and_email_user(user_id, in_app_msg, link, email_subject, email_body):
    """Vừa tạo notification in-app vừa gửi email cho 1 user."""
    if not user_id:
        return
    create_notification(user_id, in_app_msg, link)
    if email_subject:
        send_email_to_user(user_id, email_subject, email_body)


def notify_and_email_users(user_rows, in_app_msg, link, email_subject, email_body, exclude_ids=None):
    """Gửi notification + email cho danh sách user rows."""
    exclude = set(exclude_ids or [])
    sent_emails = []
    for r in user_rows:
        if r['id'] in exclude:
            continue
        create_notification(r['id'], in_app_msg, link)
        if email_subject and r['email']:
            ok, _ = send_simple_email([r['email']], email_subject, email_body)
            if ok:
                sent_emails.append(r['email'])
    return sent_emails


def find_user_by_display_name(name):
    """Tra cứu user theo full_name hoặc username (dùng cho signer_approver)."""
    if not name:
        return None
    name = name.strip()
    return get_db().execute(
        'SELECT * FROM users WHERE full_name=? OR username=? LIMIT 1',
        (name, name)
    ).fetchone()


def _get_kbc_receivers():
    """Nhân sự KBC có quyền nhận đơn (receive_kbc_order)."""
    rows = get_db().execute(
        "SELECT * FROM users WHERE organization='KBC' ORDER BY full_name, username"
    ).fetchall()
    return [r for r in rows if User(r).has_cap('receive_kbc_order')]


def _get_kbc_leaders():
    """Lãnh đạo KBC: user thuộc KBC có view_all hoặc role giám đốc/phó GĐ/admin."""
    rows = get_db().execute(
        "SELECT * FROM users WHERE organization='KBC' "
        "AND (role IN ('admin','director','deputy_director') OR permissions LIKE '%view_all%') "
        "ORDER BY full_name, username"
    ).fetchall()
    # Loại trùng + dedup theo id (LIKE có thể match nhầm)
    seen = set()
    result = []
    for r in rows:
        if r['id'] in seen:
            continue
        u = User(r)
        if u.is_admin or u.has_cap('view_all') or u.role in ('director', 'deputy_director'):
            seen.add(r['id'])
            result.append(r)
    return result


@app.context_processor
def inject_notifications():
    if not current_user.is_authenticated:
        return {}
    row = get_db().execute(
        'SELECT COUNT(*) c FROM notifications WHERE user_id=? AND is_read=0',
        (current_user.id,)
    ).fetchone()
    return {'unread_notif_count': row['c'] if row else 0}


@app.context_processor
def inject_push_config():
    return {'vapid_public_key': VAPID_PUBLIC_KEY}


# ---------- Auth ----------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        row = get_db().execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        if row and check_password_hash(row['password_hash'], password):
            login_user(User(row))
            return redirect(url_for('dashboard'))
        flash('Sai tên đăng nhập hoặc mật khẩu', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old = request.form.get('old_password', '')
        new = request.form.get('new_password', '')
        confirm = request.form.get('confirm', '')
        row = get_db().execute('SELECT password_hash FROM users WHERE id=?', (current_user.id,)).fetchone()
        if not check_password_hash(row['password_hash'], old):
            flash('Mật khẩu hiện tại không đúng', 'danger')
        elif len(new) < 6:
            flash('Mật khẩu mới tối thiểu 6 ký tự', 'danger')
        elif new != confirm:
            flash('Xác nhận mật khẩu không khớp', 'danger')
        else:
            db = get_db()
            db.execute('UPDATE users SET password_hash=? WHERE id=?',
                       (generate_password_hash(new), current_user.id))
            db.commit()
            flash('Đổi mật khẩu thành công', 'success')
            return redirect(url_for('dashboard'))
    return render_template('change_password.html')


# ---------- Users ----------
def _selected_caps(form, target_org):
    """Đọc caps được tick từ form. Lọc theo tổ chức của user đích:
    - Admin: được cấp mọi cap.
    - Manager KBC/HP89: chỉ cấp được cap của org đó + cap chung.
    """
    if current_user.is_admin:
        allowed = CAP_KEYS
    else:
        allowed = caps_for_org(target_org)
    chosen = [k for k in form.getlist('caps') if k in allowed]
    return ','.join(chosen)


def _can_manage_user(target_org, target_role=None):
    """Kiểm tra current_user có được quản lý user thuộc target_org không.
    Returns (ok: bool, error_msg: str hoặc None)."""
    if current_user.is_admin:
        return True, None
    if target_org != current_user.organization:
        return False, f'Bạn chỉ được quản lý user thuộc {current_user.organization} (cùng tổ chức)'
    if target_role == 'admin':
        return False, 'Chỉ Admin mới được quản lý tài khoản Admin'
    return True, None


def _allowed_roles_for_current():
    """Danh sách role current_user được phép gán (admin chỉ chính admin gán được)."""
    if current_user.is_admin:
        return ROLE_ORDER
    return [r for r in ROLE_ORDER if r != 'admin']


@app.route('/users')
@login_required
@manage_users_required
def users_list():
    db = get_db()
    if current_user.is_admin:
        rows = db.execute('SELECT * FROM users ORDER BY organization, created_at DESC').fetchall()
    else:
        # Manager KBC/HP89: chỉ thấy user cùng org (không thấy admin)
        rows = db.execute(
            "SELECT * FROM users WHERE organization=? AND role!='admin' ORDER BY created_at DESC",
            (current_user.organization,)
        ).fetchall()
    users = [(r, User(r)) for r in rows]
    return render_template('users.html', users=users, role_labels=ROLE_LABELS, org_labels=ORG_LABELS)


@app.route('/users/new', methods=['GET', 'POST'])
@login_required
@manage_users_required
def user_new():
    allowed_roles = _allowed_roles_for_current()
    # Non-admin: chỉ tạo user cùng org của mình. Admin: chọn org bất kỳ.
    if current_user.is_admin:
        org_choices = ORGS
    else:
        org_choices = [current_user.organization]

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip()
        role = request.form.get('role', 'staff')
        organization = request.form.get('organization', current_user.organization)

        # Non-admin: ép org = org của mình, không cho chọn admin
        if not current_user.is_admin:
            organization = current_user.organization
            if role == 'admin':
                role = 'staff'

        if organization not in ORGS:
            organization = current_user.organization

        ok, err = _can_manage_user(organization, role)
        if not ok:
            flash(err, 'danger')
            return redirect(url_for('user_new'))

        permissions = _selected_caps(request.form, organization)
        if not username or not password:
            flash('Vui lòng nhập tên đăng nhập và mật khẩu', 'danger')
        elif len(password) < 6:
            flash('Mật khẩu tối thiểu 6 ký tự', 'danger')
        elif role not in allowed_roles:
            flash('Vai trò không hợp lệ hoặc bạn không có quyền gán vai trò này', 'danger')
        else:
            try:
                db = get_db()
                db.execute(
                    'INSERT INTO users(username, password_hash, full_name, email, role, organization, permissions, created_at) VALUES (?,?,?,?,?,?,?,?)',
                    (username, generate_password_hash(password), full_name, email, role, organization, permissions,
                     datetime.now().isoformat(timespec='seconds'))
                )
                db.commit()
                flash(f'Đã tạo user "{username}" thuộc {organization}', 'success')
                return redirect(url_for('users_list'))
            except sqlite3.IntegrityError:
                flash('Tên đăng nhập đã tồn tại', 'danger')
    return render_template('user_form.html', edit_user=None, capabilities=CAPABILITIES,
                           manage_users_cap=MANAGE_USERS_CAP,
                           role_labels=ROLE_LABELS, role_order=allowed_roles,
                           orgs=org_choices, org_labels=ORG_LABELS,
                           default_caps=DEFAULT_CAPS, full_access_roles=FULL_ACCESS_ROLES,
                           cap_orgs=CAP_ORGS, force_org=(current_user.organization if not current_user.is_admin else None))


@app.route('/users/<int:uid>/edit', methods=['GET', 'POST'])
@login_required
@manage_users_required
def user_edit(uid):
    db = get_db()
    row = db.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    if not row:
        abort(404)
    eu = User(row)

    # Manager non-admin: chỉ sửa được user cùng org và không phải admin
    ok, err = _can_manage_user(eu.organization, eu.role)
    if not ok:
        flash(err, 'danger')
        return redirect(url_for('users_list'))

    allowed_roles = _allowed_roles_for_current()
    if current_user.is_admin:
        org_choices = ORGS
    else:
        org_choices = [current_user.organization]

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip()
        role = request.form.get('role', eu.role)
        organization = request.form.get('organization', eu.organization)

        # Non-admin: không cho đổi org, không cho thăng cấp admin
        if not current_user.is_admin:
            organization = eu.organization
            if role == 'admin':
                role = eu.role

        if organization not in ORGS:
            organization = eu.organization

        # Re-check sau khi normalize
        ok, err = _can_manage_user(organization, role)
        if not ok:
            flash(err, 'danger')
            return redirect(url_for('user_edit', uid=uid))

        permissions = _selected_caps(request.form, organization)
        if role not in allowed_roles:
            flash('Vai trò không hợp lệ hoặc bạn không có quyền gán vai trò này', 'danger')
        else:
            db.execute('UPDATE users SET full_name=?, email=?, role=?, organization=?, permissions=? WHERE id=?',
                       (full_name, email, role, organization, permissions, uid))
            db.commit()
            flash('Đã cập nhật user', 'success')
            return redirect(url_for('users_list'))
    return render_template('user_form.html', edit_user=row, current_caps=eu.caps,
                           capabilities=CAPABILITIES, manage_users_cap=MANAGE_USERS_CAP,
                           role_labels=ROLE_LABELS, role_order=allowed_roles,
                           orgs=org_choices, org_labels=ORG_LABELS,
                           default_caps=DEFAULT_CAPS, full_access_roles=FULL_ACCESS_ROLES,
                           cap_orgs=CAP_ORGS, force_org=(current_user.organization if not current_user.is_admin else None))


@app.route('/users/<int:uid>/delete', methods=['POST'])
@login_required
@manage_users_required
def user_delete(uid):
    if uid == current_user.id:
        flash('Không thể xoá chính mình', 'danger')
        return redirect(url_for('users_list'))
    db = get_db()
    row = db.execute('SELECT role, organization FROM users WHERE id=?', (uid,)).fetchone()
    if not row:
        abort(404)
    ok, err = _can_manage_user(row['organization'], row['role'])
    if not ok:
        flash(err, 'danger')
        return redirect(url_for('users_list'))
    db.execute('DELETE FROM users WHERE id=?', (uid,))
    db.commit()
    flash('Đã xoá user', 'success')
    return redirect(url_for('users_list'))


@app.route('/users/<int:uid>/reset', methods=['POST'])
@login_required
@manage_users_required
def user_reset(uid):
    db = get_db()
    row = db.execute('SELECT role, organization FROM users WHERE id=?', (uid,)).fetchone()
    if not row:
        abort(404)
    ok, err = _can_manage_user(row['organization'], row['role'])
    if not ok:
        flash(err, 'danger')
        return redirect(url_for('users_list'))
    new_pwd = request.form.get('new_password', '').strip()
    if len(new_pwd) < 6:
        flash('Mật khẩu mới tối thiểu 6 ký tự', 'danger')
        return redirect(url_for('users_list'))
    db.execute('UPDATE users SET password_hash=? WHERE id=?', (generate_password_hash(new_pwd), uid))
    db.commit()
    flash('Đã đặt lại mật khẩu', 'success')
    return redirect(url_for('users_list'))


# ---------- PWA ----------
@app.route('/manifest.json')
def pwa_manifest():
    resp = send_from_directory(app.static_folder, 'manifest.json')
    resp.headers['Content-Type'] = 'application/manifest+json'
    return resp


@app.route('/sw.js')
def pwa_service_worker():
    resp = send_from_directory(app.static_folder, 'sw.js')
    resp.headers['Content-Type'] = 'application/javascript'
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


# ---------- Dashboard ----------
@app.route('/')
@login_required
def dashboard():
    db = get_db()

    # Orders theo workflow
    o_ids = list_accessible_ids('order', current_user)
    def _id_filter(col_expr):
        if o_ids is None:
            return '', []
        if not o_ids:
            return ' AND 1=0', []
        placeholder = ','.join('?' * len(o_ids))
        return f' AND {col_expr} IN ({placeholder})', list(o_ids)

    f_sql, f_params = _id_filter('id')

    wf_counts = {k: 0 for k in ORDER_WORKFLOW}
    rows = db.execute(
        f"SELECT workflow_status, COUNT(*) c FROM orders WHERE 1=1{f_sql} GROUP BY workflow_status",
        f_params
    ).fetchall()
    for r in rows:
        if r['workflow_status'] in wf_counts:
            wf_counts[r['workflow_status']] = r['c']

    # Doanh thu chỉ tính đơn đã giao
    revenue_row = db.execute(
        f"SELECT COUNT(*) cnt, COALESCE(SUM(grand_total),0) revenue, "
        f"COALESCE(SUM(subtotal),0) subtotal_sum "
        f"FROM orders WHERE workflow_status='delivered_kbc'{f_sql}",
        f_params
    ).fetchone()
    o_revenue = revenue_row['revenue']
    o_delivered = revenue_row['cnt']

    # Hợp đồng quá hạn / sắp đến hạn
    c_ids = list_accessible_ids('contract', current_user)
    if c_ids is None:
        contracts = db.execute('SELECT * FROM contracts ORDER BY due_date ASC').fetchall()
    elif c_ids:
        q = f"SELECT * FROM contracts WHERE id IN ({','.join('?' * len(c_ids))}) ORDER BY due_date ASC"
        contracts = db.execute(q, list(c_ids)).fetchall()
    else:
        contracts = []
    overdue, soon = [], []
    for c in contracts:
        st, days = contract_status(c['due_date'])
        if st == 'overdue':
            overdue.append((c, days))
        elif st == 'soon':
            soon.append((c, days))
    c_count = len(contracts)

    # Media counts
    m_ids = list_accessible_ids('media', current_user)
    media_counts = {k: 0 for k in MEDIA_STATUS}
    if m_ids is None:
        mrows = db.execute('SELECT status, COUNT(*) c FROM media_posts GROUP BY status').fetchall()
    elif m_ids:
        q = (f"SELECT status, COUNT(*) c FROM media_posts WHERE id IN ({','.join('?' * len(m_ids))}) "
             f"GROUP BY status")
        mrows = db.execute(q, list(m_ids)).fetchall()
    else:
        mrows = []
    for r in mrows:
        if r['status'] in media_counts:
            media_counts[r['status']] = r['c']

    return render_template('dashboard.html',
                           wf_counts=wf_counts,
                           o_revenue=o_revenue,
                           o_delivered=o_delivered,
                           overdue=overdue, soon=soon, c_count=c_count,
                           media_counts=media_counts)


# ============================================================
# ---------- ORDERS (Đặt Hàng) ----------
# ============================================================
def _parse_order_items(form):
    """Đọc các dòng item từ form mới: chỉ cần product_id + qty (nguyên).
    Các thông tin khác (tên SP, ĐVT, quy cách, đơn giá) lấy tự động từ
    bảng products do KBC tạo. KHÔNG còn chiết khấu."""
    items = []
    prod_ids = form.getlist('item_product_id[]')
    qtys = form.getlist('item_qty[]')
    db = get_db()
    for i, pid_str in enumerate(prod_ids):
        pid_str = (pid_str or '').strip()
        if not pid_str:
            continue
        try:
            pid = int(pid_str)
        except (ValueError, TypeError):
            continue
        p = db.execute('SELECT * FROM products WHERE id=?', (pid,)).fetchone()
        if not p:
            continue
        # SL là số nguyên — bỏ phần thập phân nếu user nhập
        try:
            qty_raw = (qtys[i] if i < len(qtys) else '0') or '0'
            qty = int(float(str(qty_raw).replace(',', '').strip()))
        except (ValueError, AttributeError):
            qty = 0
        if qty <= 0:
            continue
        price = float(p['default_price'] or 0)
        amount = qty * price
        items.append({
            'product_id': pid,
            'product_name': p['name'],
            'unit': p['unit'] or '',
            'packaging': p['packaging'] or '',
            'qty': qty,
            'unit_price': price,
            'amount': amount,
            'discount': 0,          # bỏ chiết khấu
            'amount_after': amount, # = amount vì không có CK
            'note': '',
            'sort_order': i,
        })
    return items


def _compute_order_totals(items, vat_percent):
    subtotal = sum(it['amount_after'] for it in items)
    vat_amount = subtotal * (vat_percent / 100.0)
    grand_total = subtotal + vat_amount
    return subtotal, vat_amount, grand_total


@app.route('/orders')
@login_required
def orders_list():
    db = get_db()
    status_filter = request.args.get('status', '').strip()
    base_q = '''SELECT o.*, u.username owner_user, u.full_name owner_name, u.organization owner_org
                FROM orders o JOIN users u ON o.owner_id=u.id'''

    if current_user.can_see_all:
        # Admin / view_all: thấy tất cả
        conditions, params = [], []
        if status_filter and status_filter in ORDER_WORKFLOW:
            conditions.append('o.workflow_status=?')
            params.append(status_filter)
        where = (' WHERE ' + ' AND '.join(conditions)) if conditions else ''
        rows = db.execute(base_q + where + ' ORDER BY o.order_date DESC, o.id DESC', params).fetchall()

    elif current_user.is_kbc:
        # KBC user: chỉ thấy đơn từ approved_hp89 trở đi (HP89 đã duyệt)
        kbc_statuses = ('approved_hp89', 'received_kbc', 'delivered_kbc')
        conditions = ['o.workflow_status IN ({})'.format(','.join('?' * len(kbc_statuses)))]
        params = list(kbc_statuses)
        if status_filter and status_filter in kbc_statuses:
            conditions = ['o.workflow_status=?']
            params = [status_filter]
        elif status_filter and status_filter in ORDER_WORKFLOW:
            # nếu filter không phải KBC statuses thì không trả kết quả
            conditions.append('1=0')
        where = ' WHERE ' + ' AND '.join(conditions)
        rows = db.execute(base_q + where + ' ORDER BY o.order_date DESC, o.id DESC', params).fetchall()

    else:
        # HP89 user: chỉ thấy đơn mình sở hữu + được chia sẻ
        ids = list_accessible_ids('order', current_user)
        if ids is None:
            conditions, params = [], []
            if status_filter and status_filter in ORDER_WORKFLOW:
                conditions.append('o.workflow_status=?')
                params.append(status_filter)
            where = (' WHERE ' + ' AND '.join(conditions)) if conditions else ''
            rows = db.execute(base_q + where + ' ORDER BY o.order_date DESC, o.id DESC', params).fetchall()
        elif ids:
            placeholder = ','.join('?' * len(ids))
            conditions = [f'o.id IN ({placeholder})']
            params = list(ids)
            if status_filter and status_filter in ORDER_WORKFLOW:
                conditions.append('o.workflow_status=?')
                params.append(status_filter)
            where = ' WHERE ' + ' AND '.join(conditions)
            rows = db.execute(base_q + where + ' ORDER BY o.order_date DESC, o.id DESC', params).fetchall()
        else:
            rows = []

    return render_template('orders_list.html', orders=rows, current_status=status_filter)


def _get_hp89_approvers():
    """Trả về list user thuộc HP89 (hoặc admin) có quyền duyệt đơn HP89."""
    rows = get_db().execute(
        "SELECT * FROM users WHERE organization='HP89' OR role='admin' "
        "ORDER BY full_name, username"
    ).fetchall()
    return [r for r in rows if User(r).has_cap('approve_hp89_order')]


@app.route('/orders/new', methods=['GET', 'POST'])
@login_required
@require_cap('create_order')
def order_new():
    db = get_db()
    if request.method == 'POST':
        return _save_order(None)
    products = db.execute('SELECT * FROM products WHERE active=1 ORDER BY name').fetchall()
    invoice_entities = db.execute('SELECT * FROM invoice_entities ORDER BY company_name').fetchall()
    hp89_approvers = _get_hp89_approvers()
    return render_template('order_form.html', order=None, items=[], products=products,
                           invoice_entities=invoice_entities, hp89_approvers=hp89_approvers)


def _save_order(order_id):
    db = get_db()
    is_new = order_id is None
    code = request.form.get('code', '').strip()
    customer_name = request.form.get('customer_name', '').strip()
    if not customer_name:
        flash('Vui lòng nhập tên nơi nhận hàng', 'danger')
        return redirect(url_for('order_new') if not order_id else url_for('order_edit', oid=order_id))
    try:
        vat_percent = float(request.form.get('vat_percent') or 8)
    except ValueError:
        vat_percent = 8.0
    items = _parse_order_items(request.form)
    subtotal, vat_amount, grand_total = _compute_order_totals(items, vat_percent)
    paid_amount = _parse_money(request.form.get('paid_amount'))

    of, _ = save_upload(request.files.get('order_file'), ORDER_DIR)
    inv, _ = save_upload(request.files.get('invoice_file'), ORDER_DIR)
    dlv, _ = save_upload(request.files.get('delivery_file'), ORDER_DIR)
    oth, _ = save_upload(request.files.get('other_file'), ORDER_DIR)
    wh, _ = save_upload(request.files.get('warehouse_file'), ORDER_DIR)

    now = datetime.now().isoformat(timespec='seconds')
    fields = (code,
              request.form.get('order_date') or None,
              customer_name,
              request.form.get('customer_phone', '').strip(),
              request.form.get('customer_address', '').strip(),
              request.form.get('invoice_company', '').strip(),
              request.form.get('invoice_tax_code', '').strip(),
              request.form.get('invoice_address', '').strip(),
              request.form.get('invoice_email', '').strip(),
              request.form.get('invoice_phone', '').strip(),
              request.form.get('payment_method', '').strip(),
              request.form.get('bank_account', '').strip(),
              request.form.get('signer_creator', '').strip(),
              request.form.get('signer_approver', '').strip(),
              request.form.get('signer_deliverer', '').strip(),
              subtotal, vat_percent, vat_amount, grand_total,
              request.form.get('notes', '').strip(),
              paid_amount)

    if order_id is None:
        cur = db.execute('''INSERT INTO orders
            (code, order_date, customer_name, customer_phone, customer_address,
             invoice_company, invoice_tax_code, invoice_address, invoice_email, invoice_phone,
             payment_method, bank_account, signer_creator, signer_approver, signer_deliverer,
             subtotal, vat_percent, vat_amount, grand_total, notes,
             paid_amount,
             order_file, invoice_file, delivery_file, other_file, warehouse_file,
             workflow_status, owner_id, created_at, updated_at)
            VALUES (?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?, ?, ?,?,?,?,?, ?,?,?,?)''',
            fields + (of, inv, dlv, oth, wh, 'draft', current_user.id, now, now))
        order_id = cur.lastrowid
    else:
        old = db.execute('SELECT order_file, invoice_file, delivery_file, other_file, warehouse_file FROM orders WHERE id=?',
                         (order_id,)).fetchone()
        of = of or old['order_file']
        inv = inv or old['invoice_file']
        dlv = dlv or old['delivery_file']
        oth = oth or old['other_file']
        wh = wh or old['warehouse_file']
        db.execute('''UPDATE orders SET
            code=?, order_date=?, customer_name=?, customer_phone=?, customer_address=?,
            invoice_company=?, invoice_tax_code=?, invoice_address=?, invoice_email=?, invoice_phone=?,
            payment_method=?, bank_account=?, signer_creator=?, signer_approver=?, signer_deliverer=?,
            subtotal=?, vat_percent=?, vat_amount=?, grand_total=?, notes=?,
            paid_amount=?,
            order_file=?, invoice_file=?, delivery_file=?, other_file=?, warehouse_file=?, updated_at=?
            WHERE id=?''',
            fields + (of, inv, dlv, oth, wh, now, order_id))
        db.execute('DELETE FROM order_items WHERE order_id=?', (order_id,))

    for it in items:
        db.execute('''INSERT INTO order_items
            (order_id, product_id, product_name, unit, packaging,
             qty, unit_price, amount, discount, amount_after, note, sort_order)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (order_id, it['product_id'], it['product_name'], it['unit'], it['packaging'],
             it['qty'], it['unit_price'], it['amount'], it['discount'], it['amount_after'],
             it['note'], it['sort_order']))

    if is_new:
        link = url_for('order_view', oid=order_id)
        notify_cap_users('notify_order',
                         f'[ĐƠN HÀNG MỚI] {code or customer_name} đã được HP89 tạo (nháp)',
                         link, exclude_ids=[current_user.id])
    db.commit()

    # Nếu user bấm "Lưu & Gửi Lãnh đạo HP89 duyệt" thì transition ngay
    action = request.form.get('action', 'save')
    if action == 'save_and_submit':
        # Chỉ submit được khi đang ở trạng thái draft (mới tạo hoặc edit nháp)
        ok, msg = _do_submit_order(order_id)
        if ok:
            flash('Đã lưu & GỬI Lãnh đạo HP89 duyệt — email đã được gửi', 'success')
        else:
            flash(f'Đã lưu đơn nhưng không gửi được: {msg}', 'warning')
        return redirect(url_for('order_view', oid=order_id))

    flash('Đã lưu đơn hàng', 'success')
    return redirect(url_for('order_view', oid=order_id))


@app.route('/orders/<int:oid>')
@login_required
def order_view(oid):
    # KBC users can view orders in approved/received/delivered status without explicit share
    if current_user.is_kbc and not current_user.can_see_all:
        o_check = get_db().execute('SELECT workflow_status FROM orders WHERE id=?', (oid,)).fetchone()
        if o_check and o_check['workflow_status'] not in ('approved_hp89', 'received_kbc', 'delivered_kbc'):
            abort(403)
        elif not o_check:
            abort(404)
    elif not can_view('order', oid, current_user):
        abort(403)
    db = get_db()
    o = db.execute('''SELECT o.*, u.username owner_user, u.full_name owner_name, u.organization owner_org,
                             a.full_name approver_name, a.username approver_user,
                             r.full_name receiver_name, r.username receiver_user,
                             d.full_name deliverer_name, d.username deliverer_user
                      FROM orders o JOIN users u ON o.owner_id=u.id
                      LEFT JOIN users a ON o.approved_by=a.id
                      LEFT JOIN users r ON o.received_by=r.id
                      LEFT JOIN users d ON o.delivered_by=d.id
                      WHERE o.id=?''', (oid,)).fetchone()
    if not o:
        abort(404)
    items = db.execute('SELECT * FROM order_items WHERE order_id=? ORDER BY sort_order, id', (oid,)).fetchall()
    comments = get_comments('order', oid)
    return render_template('order_view.html', o=o, items=items, comments=comments,
                           editable=can_edit('order', oid, current_user))


@app.route('/orders/<int:oid>/edit', methods=['GET', 'POST'])
@login_required
def order_edit(oid):
    if not can_edit('order', oid, current_user):
        abort(403)
    db = get_db()
    o = db.execute('SELECT * FROM orders WHERE id=?', (oid,)).fetchone()
    if not o:
        abort(404)
    # Chỉ cho sửa khi còn ở trạng thái draft hoặc admin
    if o['workflow_status'] not in ('draft',) and not current_user.is_admin:
        flash('Chỉ sửa được đơn ở trạng thái Nháp. Liên hệ Admin nếu cần điều chỉnh đơn đã gửi duyệt.', 'warning')
        return redirect(url_for('order_view', oid=oid))
    if request.method == 'POST':
        return _save_order(oid)
    items = db.execute('SELECT * FROM order_items WHERE order_id=? ORDER BY sort_order, id', (oid,)).fetchall()
    products = db.execute('SELECT * FROM products WHERE active=1 ORDER BY name').fetchall()
    invoice_entities = db.execute('SELECT * FROM invoice_entities ORDER BY company_name').fetchall()
    hp89_approvers = _get_hp89_approvers()
    return render_template('order_form.html', order=o, items=items, products=products,
                           invoice_entities=invoice_entities, hp89_approvers=hp89_approvers)


@app.route('/orders/<int:oid>/delete', methods=['POST'])
@login_required
def order_delete(oid):
    if not can_edit('order', oid, current_user):
        abort(403)
    db = get_db()
    o = db.execute('SELECT owner_id, order_file, invoice_file, delivery_file, other_file, warehouse_file FROM orders WHERE id=?',
                   (oid,)).fetchone()
    if not o:
        abort(404)
    if not current_user.full_access and o['owner_id'] != current_user.id:
        flash('Chỉ chủ đơn hoặc Admin mới được xoá', 'danger')
        return redirect(url_for('order_view', oid=oid))
    db.execute('DELETE FROM orders WHERE id=?', (oid,))
    db.commit()
    for f in (o['order_file'], o['invoice_file'], o['delivery_file'], o['other_file'], o['warehouse_file']):
        if f:
            try:
                os.remove(os.path.join(ORDER_DIR, f))
            except OSError:
                pass
    flash('Đã xoá đơn hàng', 'success')
    return redirect(url_for('orders_list'))


@app.route('/orders/<int:oid>/file/<which>')
@login_required
def order_file(oid, which):
    if which not in ('order_file', 'invoice_file', 'delivery_file', 'other_file', 'warehouse_file'):
        abort(400)
    if not can_view('order', oid, current_user):
        abort(403)
    o = get_db().execute(f'SELECT {which} AS f FROM orders WHERE id=?', (oid,)).fetchone()
    if not o or not o['f']:
        abort(404)
    return send_from_directory(ORDER_DIR, o['f'], as_attachment=True)


# ---- Order workflow transitions ----
def _order_or_404(oid):
    db = get_db()
    o = db.execute('SELECT * FROM orders WHERE id=?', (oid,)).fetchone()
    if not o:
        abort(404)
    return o


@app.route('/orders/<int:oid>/submit', methods=['POST'])
@login_required
def _do_submit_order(oid):
    """Logic chuyển đơn từ Nháp -> Chờ HP89 duyệt + gửi notification + email."""
    db = get_db()
    o = db.execute('SELECT * FROM orders WHERE id=?', (oid,)).fetchone()
    if not o:
        return False, 'Đơn không tồn tại'
    if o['workflow_status'] != 'draft':
        return False, 'Đơn không ở trạng thái Nháp'
    now = datetime.now().isoformat(timespec='seconds')
    db.execute('UPDATE orders SET workflow_status=?, submitted_at=?, updated_at=? WHERE id=?',
               ('pending_hp89', now, now, oid))
    link = url_for('order_view', oid=oid)
    abs_link = absolute_link(link)
    code_or_name = o["code"] or o["customer_name"]
    creator_name = current_user.full_name or current_user.username

    # 1) Email + thông báo cho người duyệt được chọn cụ thể
    approver_row = find_user_by_display_name(o['signer_approver'])
    if approver_row:
        subject = f'[KBC-HP89] Đơn hàng "{code_or_name}" cần bạn duyệt'
        body = (
            f'Kính gửi {approver_row["full_name"] or approver_row["username"]},\n\n'
            f'Bạn được chọn làm người duyệt cho đơn hàng:\n'
            f'  • Mã đơn: {o["code"] or "(chưa có)"}\n'
            f'  • Nơi nhận: {o["customer_name"]}\n'
            f'  • Tổng cộng: {money0_fmt(o["grand_total"])} đ\n'
            f'  • Người tạo: {creator_name}\n\n'
            f'Vui lòng đăng nhập hệ thống để duyệt: {abs_link}\n\n'
            f'Trân trọng,\nKBC-HP89'
        )
        notify_and_email_user(
            approver_row['id'],
            f'[CHỜ BẠN DUYỆT] Đơn "{code_or_name}" cần duyệt',
            link, subject, body
        )

    # 2) Thông báo (không email) cho các lãnh đạo HP89 khác có cap approve_hp89_order
    exclude = [current_user.id]
    if approver_row:
        exclude.append(approver_row['id'])
    notify_cap_users('approve_hp89_order',
                     f'[CHỜ DUYỆT] Đơn "{code_or_name}" cần Lãnh đạo HP89 duyệt',
                     link, exclude_ids=exclude, organization='HP89')
    db.commit()
    return True, 'OK'


def money0_fmt(v):
    """Format tiền VN dùng cho email (không dùng filter Jinja)."""
    try:
        return f'{float(v or 0):,.0f}'.replace(',', '.')
    except (ValueError, TypeError):
        return str(v)


@app.route('/orders/<int:oid>/submit', methods=['POST'])
@login_required
def order_submit(oid):
    """HP89 gửi đơn nháp đi để lãnh đạo HP89 duyệt."""
    o = _order_or_404(oid)
    if not (current_user.is_admin or o['owner_id'] == current_user.id):
        abort(403)
    ok, msg = _do_submit_order(oid)
    if ok:
        flash('Đã gửi đơn cho Lãnh đạo HP89 duyệt — email đã được gửi đến người duyệt', 'success')
    else:
        flash(msg, 'warning')
    return redirect(url_for('order_view', oid=oid))


@app.route('/orders/<int:oid>/approve-hp89', methods=['POST'])
@login_required
@require_cap('approve_hp89_order')
def order_approve_hp89(oid):
    """Lãnh đạo HP89 duyệt -> đẩy sang KBC. Thông báo + email cho KBC."""
    o = _order_or_404(oid)
    if o['workflow_status'] != 'pending_hp89':
        flash('Đơn không ở trạng thái Chờ HP89 duyệt', 'warning')
        return redirect(url_for('order_view', oid=oid))
    db = get_db()
    now = datetime.now().isoformat(timespec='seconds')
    db.execute('''UPDATE orders SET workflow_status=?, approved_by=?, approved_at=?, updated_at=?
                  WHERE id=?''',
               ('approved_hp89', current_user.id, now, now, oid))
    link = url_for('order_view', oid=oid)
    abs_link = absolute_link(link)
    code_or_name = o["code"] or o["customer_name"]
    approver_name = current_user.full_name or current_user.username

    # 1) Báo (in-app only) cho người tạo đơn HP89
    notify_user(o['owner_id'],
                f'Lãnh đạo HP89 đã DUYỆT đơn "{code_or_name}"', link)

    # 2) Email + thông báo cho nhân sự KBC tiếp nhận đơn
    kbc_recv = _get_kbc_receivers()
    subj_recv = f'[KBC TIẾP NHẬN] Đơn "{code_or_name}" đã được HP89 duyệt'
    body_recv = (
        f'Kính gửi anh/chị,\n\n'
        f'Đơn hàng từ HP89 vừa được Lãnh đạo HP89 duyệt và chuyển sang KBC tiếp nhận:\n'
        f'  • Mã đơn: {o["code"] or "(chưa có)"}\n'
        f'  • Nơi nhận: {o["customer_name"]}\n'
        f'  • SĐT: {o["customer_phone"] or "—"}\n'
        f'  • Tổng cộng: {money0_fmt(o["grand_total"])} đ\n'
        f'  • Người duyệt HP89: {approver_name}\n\n'
        f'Vui lòng đăng nhập xác nhận tiếp nhận đơn: {abs_link}\n\n'
        f'Trân trọng,\nKBC-HP89'
    )
    notify_and_email_users(kbc_recv,
                           f'[KBC NHẬN ĐƠN] HP89 đã duyệt đơn "{code_or_name}" — vui lòng tiếp nhận',
                           link, subj_recv, body_recv,
                           exclude_ids=[current_user.id])

    # 3) Email + thông báo cho lãnh đạo KBC
    kbc_lead = _get_kbc_leaders()
    recv_ids = {r['id'] for r in kbc_recv}
    subj_lead = f'[Lãnh đạo KBC] Đơn "{code_or_name}" đã được HP89 duyệt'
    body_lead = (
        f'Đơn hàng từ HP89 đã được Lãnh đạo HP89 duyệt, đang chờ KBC tiếp nhận:\n'
        f'  • Mã đơn: {o["code"] or "(chưa có)"}\n'
        f'  • Nơi nhận: {o["customer_name"]}\n'
        f'  • Tổng cộng: {money0_fmt(o["grand_total"])} đ\n'
        f'  • Duyệt bởi: {approver_name}\n\n'
        f'Xem chi tiết: {abs_link}\n'
    )
    # Lọc bỏ những người đã được gửi ở mục KBC tiếp nhận
    leaders_only = [r for r in kbc_lead if r['id'] not in recv_ids]
    notify_and_email_users(leaders_only,
                           f'HP89 đã duyệt đơn "{code_or_name}" — đang chờ KBC tiếp nhận',
                           link, subj_lead, body_lead,
                           exclude_ids=[current_user.id])

    db.commit()
    flash('Đã duyệt đơn — đẩy sang KBC tiếp nhận, đã gửi email + thông báo', 'success')
    return redirect(url_for('order_view', oid=oid))


@app.route('/orders/<int:oid>/reject-hp89', methods=['POST'])
@login_required
@require_cap('approve_hp89_order')
def order_reject_hp89(oid):
    """Lãnh đạo HP89 trả lại đơn về nháp để chỉnh sửa."""
    o = _order_or_404(oid)
    if o['workflow_status'] != 'pending_hp89':
        flash('Đơn không ở trạng thái Chờ HP89 duyệt', 'warning')
        return redirect(url_for('order_view', oid=oid))
    db = get_db()
    now = datetime.now().isoformat(timespec='seconds')
    note = request.form.get('note', '').strip()
    db.execute('UPDATE orders SET workflow_status=?, updated_at=? WHERE id=?',
               ('draft', now, oid))
    link = url_for('order_view', oid=oid)
    msg = f'Lãnh đạo HP89 đã TRẢ LẠI đơn "{o["code"] or o["customer_name"]}"'
    if note:
        msg += f' — Lý do: {note}'
    notify_user(o['owner_id'], msg, link)
    db.commit()
    flash('Đã trả lại đơn về Nháp', 'warning')
    return redirect(url_for('order_view', oid=oid))


@app.route('/orders/<int:oid>/receive-kbc', methods=['POST'])
@login_required
@require_cap('receive_kbc_order')
def order_receive_kbc(oid):
    """Nhân viên KBC nhận đơn."""
    o = _order_or_404(oid)
    if o['workflow_status'] != 'approved_hp89':
        flash('Đơn chưa được Lãnh đạo HP89 duyệt, chưa thể nhận', 'warning')
        return redirect(url_for('order_view', oid=oid))
    db = get_db()
    now = datetime.now().isoformat(timespec='seconds')
    db.execute('''UPDATE orders SET workflow_status=?, received_by=?, received_at=?, updated_at=?
                  WHERE id=?''',
               ('received_kbc', current_user.id, now, now, oid))
    link = url_for('order_view', oid=oid)
    notify_user(o['owner_id'],
                f'KBC đã NHẬN đơn "{o["code"] or o["customer_name"]}" — đang chuẩn bị giao hàng', link)
    if o['approved_by']:
        notify_user(o['approved_by'],
                    f'KBC đã NHẬN đơn "{o["code"] or o["customer_name"]}"', link)
    db.commit()
    flash('Đã tiếp nhận đơn hàng — chuẩn bị giao hàng', 'success')
    return redirect(url_for('order_view', oid=oid))


@app.route('/orders/<int:oid>/deliver-kbc', methods=['POST'])
@login_required
@require_cap('deliver_kbc_order')
def order_deliver_kbc(oid):
    """Nhân viên KBC tick đã giao. Thông báo + email cho HP89 owner, HP89 approver, lãnh đạo KBC."""
    o = _order_or_404(oid)
    if o['workflow_status'] != 'received_kbc':
        flash('Đơn chưa ở trạng thái KBC đã nhận, chưa thể tick giao', 'warning')
        return redirect(url_for('order_view', oid=oid))
    db = get_db()
    now = datetime.now().isoformat(timespec='seconds')
    db.execute('''UPDATE orders SET workflow_status=?, delivered_by=?, delivered_at=?, updated_at=?
                  WHERE id=?''',
               ('delivered_kbc', current_user.id, now, now, oid))
    link = url_for('order_view', oid=oid)
    abs_link = absolute_link(link)
    code_or_name = o["code"] or o["customer_name"]
    deliverer_name = current_user.full_name or current_user.username

    subject = f'[Đã giao] Đơn "{code_or_name}" đã hoàn thành'
    body_base = (
        f'KBC đã giao xong đơn hàng:\n'
        f'  • Mã đơn: {o["code"] or "(chưa có)"}\n'
        f'  • Nơi nhận: {o["customer_name"]}\n'
        f'  • Tổng cộng: {money0_fmt(o["grand_total"])} đ\n'
        f'  • Người giao: {deliverer_name}\n\n'
        f'Xem chi tiết: {abs_link}\n'
    )
    msg_inapp = f'KBC đã GIAO XONG đơn "{code_or_name}"'

    sent_to_ids = set()

    # 1) HP89 — người tạo đơn (owner)
    if o['owner_id']:
        notify_and_email_user(o['owner_id'], msg_inapp, link, subject,
                              f'Kính gửi anh/chị,\n\n{body_base}\nTrân trọng,\nKBC-HP89')
        sent_to_ids.add(o['owner_id'])

    # 2) HP89 — Lãnh đạo HP89 đã duyệt đơn
    if o['approved_by'] and o['approved_by'] not in sent_to_ids:
        notify_and_email_user(o['approved_by'], msg_inapp, link, subject,
                              f'Kính gửi Lãnh đạo,\n\nĐơn anh/chị đã duyệt nay hoàn thành:\n{body_base}')
        sent_to_ids.add(o['approved_by'])

    # 3) Lãnh đạo KBC
    leaders = _get_kbc_leaders()
    leaders_email_body = f'Đơn hàng đã hoàn thành giao:\n{body_base}'
    notify_and_email_users(leaders, msg_inapp, link, subject, leaders_email_body,
                           exclude_ids=list(sent_to_ids) + [current_user.id])

    # 4) Nhân viên KBC đã nhận đơn (nếu khác người giao) — in-app only
    if o['received_by'] and o['received_by'] != current_user.id and o['received_by'] not in sent_to_ids:
        notify_user(o['received_by'], f'Đơn "{code_or_name}" đã được giao xong', link)

    db.commit()
    flash('Đã tick GIAO XONG — đơn hoàn thành, đã gửi email + thông báo', 'success')
    return redirect(url_for('order_view', oid=oid))


@app.route('/orders/<int:oid>/reopen', methods=['POST'])
@login_required
@admin_required
def order_reopen(oid):
    """Admin mở lại đơn về trạng thái Nháp (trường hợp cần điều chỉnh)."""
    o = _order_or_404(oid)
    db = get_db()
    now = datetime.now().isoformat(timespec='seconds')
    db.execute('''UPDATE orders SET workflow_status='draft', approved_by=NULL, approved_at=NULL,
                  received_by=NULL, received_at=NULL, delivered_by=NULL, delivered_at=NULL,
                  submitted_at=NULL, updated_at=? WHERE id=?''', (now, oid))
    db.commit()
    flash('Admin đã mở lại đơn về trạng thái Nháp', 'warning')
    return redirect(url_for('order_view', oid=oid))


# ---------- Order print ----------
def _num_to_vn_words(n):
    if n is None:
        return ''
    n = int(round(n))
    if n == 0:
        return 'Không đồng'
    digits = ['không', 'một', 'hai', 'ba', 'bốn', 'năm', 'sáu', 'bảy', 'tám', 'chín']

    def read_three(num, full=False):
        tram, chuc, donvi = num // 100, (num // 10) % 10, num % 10
        parts = []
        if tram > 0:
            parts.append(f'{digits[tram]} trăm')
        elif full and (chuc > 0 or donvi > 0):
            parts.append('không trăm')
        if chuc > 1:
            parts.append(f'{digits[chuc]} mươi')
            if donvi == 1:
                parts.append('mốt')
            elif donvi == 5:
                parts.append('lăm')
            elif donvi > 0:
                parts.append(digits[donvi])
        elif chuc == 1:
            parts.append('mười')
            if donvi == 5:
                parts.append('lăm')
            elif donvi > 0:
                parts.append(digits[donvi])
        elif chuc == 0 and donvi > 0:
            if tram > 0 or full:
                parts.append('lẻ')
            parts.append(digits[donvi])
        return ' '.join(parts)

    units = ['', ' nghìn', ' triệu', ' tỷ']
    groups = []
    i = 0
    while n > 0:
        groups.append((n % 1000, units[i] if i < len(units) else ''))
        n //= 1000
        i += 1
    groups.reverse()
    is_first = True
    result_parts = []
    for num, unit in groups:
        if num == 0:
            continue
        text = read_three(num, full=not is_first)
        result_parts.append(text + unit)
        is_first = False
    s = ' '.join(result_parts).strip()
    if not s:
        return 'Không đồng'
    return s[0].upper() + s[1:] + ' đồng'


app.jinja_env.globals['num_to_vn_words'] = _num_to_vn_words


@app.route('/orders/<int:oid>/print')
@login_required
def order_print(oid):
    if not can_view('order', oid, current_user):
        abort(403)
    db = get_db()
    o = db.execute('''SELECT o.*, u.username owner_user, u.full_name owner_name,
                             a.full_name approver_name, a.username approver_user
                      FROM orders o
                      JOIN users u ON o.owner_id=u.id
                      LEFT JOIN users a ON o.approved_by=a.id
                      WHERE o.id=?''', (oid,)).fetchone()
    if not o:
        abort(404)
    # Cho phép in ở mọi trạng thái (kể cả Nháp) — nhân viên HP89 có thể in để
    # mang phiếu đi xin chữ ký nội bộ trước khi gửi duyệt.
    items = db.execute('SELECT * FROM order_items WHERE order_id=? ORDER BY sort_order, id', (oid,)).fetchall()
    total_discount = sum((it['amount'] or 0) - (it['amount_after'] or 0) for it in items)
    return render_template('order_print.html', o=o, items=items, total_discount=total_discount)


# ============================================================
# ---------- CONTRACTS (Hợp Đồng KBC-HP89) ----------
# ============================================================
DOC_CATEGORIES = [
    ('hop_dong', 'Hợp đồng'),
    ('phu_luc', 'Phụ lục'),
    ('bien_ban_thoa_thuan', 'Biên bản thỏa thuận'),
    ('bien_ban_ban_giao', 'Biên bản bàn giao'),
    ('bien_ban_nghiem_thu', 'Biên bản nghiệm thu'),
    ('hoa_don_vat', 'Hóa đơn VAT'),
    ('chung_tu_khac', 'File chứng từ khác'),
]
DOC_CATEGORY_KEYS = {k for k, _ in DOC_CATEGORIES}
app.jinja_env.globals['DOC_CATEGORIES'] = DOC_CATEGORIES


def get_attachments(record_type, record_id):
    rows = get_db().execute(
        'SELECT * FROM attachments WHERE record_type=? AND record_id=? ORDER BY category, id',
        (record_type, record_id)).fetchall()
    grouped = {k: [] for k, _ in DOC_CATEGORIES}
    for r in rows:
        grouped.setdefault(r['category'], []).append(r)
    return grouped


def _receiver_user_list():
    return get_db().execute(
        'SELECT id, full_name, username, email, organization FROM users WHERE id != ? ORDER BY organization, full_name, username',
        (current_user.id,)).fetchall()


def receivers_from_form(form):
    ids = []
    for v in form.getlist('receiver_ids'):
        try:
            ids.append(int(v))
        except (TypeError, ValueError):
            continue
    if not ids:
        return [], '', None
    db = get_db()
    q = f"SELECT id, full_name, username, email FROM users WHERE id IN ({','.join('?' * len(ids))})"
    rows = db.execute(q, ids).fetchall()
    by_id = {r['id']: r for r in rows}
    names, emails, valid_ids = [], [], []
    for uid in ids:
        r = by_id.get(uid)
        if not r:
            continue
        valid_ids.append(uid)
        names.append(r['full_name'] or r['username'])
        if r['email']:
            emails.append(r['email'])
    return valid_ids, '; '.join(names), (','.join(emails) or None)


@app.route('/contracts')
@login_required
def contracts_list():
    db = get_db()
    ids = list_accessible_ids('contract', current_user)
    if ids is None:
        rows = db.execute('''SELECT c.*, u.full_name owner_name, u.username owner_user, u.organization owner_org
                             FROM contracts c JOIN users u ON c.owner_id=u.id
                             ORDER BY c.due_date ASC''').fetchall()
    elif ids:
        q = f'''SELECT c.*, u.full_name owner_name, u.username owner_user, u.organization owner_org
                FROM contracts c JOIN users u ON c.owner_id=u.id
                WHERE c.id IN ({','.join('?' * len(ids))}) ORDER BY c.due_date ASC'''
        rows = db.execute(q, list(ids)).fetchall()
    else:
        rows = []
    return render_template('contracts_list.html', contracts=rows)


@app.route('/contracts/new', methods=['GET', 'POST'])
@login_required
@require_cap('manage_contract')
def contract_new():
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        if not title:
            flash('Tên hợp đồng không được trống', 'danger')
            return redirect(url_for('contract_new'))
        cf, _ = save_upload(request.files.get('contract_file'), CONTRACT_DIR)
        hf, _ = save_upload(request.files.get('handover_file'), CONTRACT_DIR)
        af, _ = save_upload(request.files.get('appendix_file'), CONTRACT_DIR)
        inv, _ = save_upload(request.files.get('invoice_file'), CONTRACT_DIR)
        oth, _ = save_upload(request.files.get('other_file'), CONTRACT_DIR)
        try:
            pct = int(request.form.get('progress_percent') or 0)
            pct = max(0, min(100, pct))
        except ValueError:
            pct = 0
        rcv_ids, rcv_names, rcv_emails = receivers_from_form(request.form)
        now = datetime.now().isoformat(timespec='seconds')
        db = get_db()
        cur = db.execute('''INSERT INTO contracts
            (code, title, partner, supplier_tax, contact_person, contact_phone,
             receiver, receiver_email, receiver_ids, total_value, paid_amount,
             contract_date, due_date, progress, progress_percent, notes,
             contract_file, handover_file, appendix_file, invoice_file, other_file,
             owner_id, created_at, updated_at)
            VALUES (?,?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?, ?,?,?)''',
            (request.form.get('code', '').strip(),
             title,
             request.form.get('partner', '').strip(),
             request.form.get('supplier_tax', '').strip(),
             request.form.get('contact_person', '').strip(),
             request.form.get('contact_phone', '').strip(),
             rcv_names, rcv_emails,
             ','.join(str(i) for i in rcv_ids) or None,
             _parse_money(request.form.get('total_value')),
             _parse_money(request.form.get('paid_amount')),
             request.form.get('contract_date') or None,
             request.form.get('due_date') or None,
             request.form.get('progress', '').strip(),
             pct,
             request.form.get('notes', '').strip(),
             cf, hf, af, inv, oth, current_user.id, now, now))
        new_id = cur.lastrowid
        link = url_for('contract_view', cid=new_id)
        for uid in rcv_ids:
            if uid != current_user.id:
                create_notification(uid, f'Bạn là người nhận hợp đồng: "{title}"', link)
        notify_cap_users('notify_contract', f'Có hợp đồng mới: "{title}"', link,
                         exclude_ids=[current_user.id] + rcv_ids)
        db.commit()
        flash('Đã tạo hợp đồng', 'success')
        return redirect(link)
    return render_template('contract_form.html', contract=None,
                           users=_receiver_user_list(), selected_receiver_ids=[])


@app.route('/contracts/<int:cid>')
@login_required
def contract_view(cid):
    if not can_view('contract', cid, current_user):
        abort(403)
    c = get_db().execute('''SELECT c.*, u.full_name owner_name, u.username owner_user, u.organization owner_org
                            FROM contracts c JOIN users u ON c.owner_id=u.id WHERE c.id=?''',
                         (cid,)).fetchone()
    if not c:
        abort(404)
    shared = get_db().execute('''SELECT rp.perm, u.id, u.username, u.full_name, u.organization
                                 FROM record_permissions rp JOIN users u ON u.id=rp.user_id
                                 WHERE rp.record_type=? AND rp.record_id=?''',
                              ('contract', cid)).fetchall()
    all_users = get_db().execute('SELECT id, username, full_name, organization FROM users WHERE id != ?',
                                 (c['owner_id'],)).fetchall()
    comments = get_comments('contract', cid)
    attachments = get_attachments('contract', cid)
    return render_template('contract_view.html', c=c, shared=shared, all_users=all_users,
                           comments=comments, attachments=attachments,
                           editable=can_edit('contract', cid, current_user))


@app.route('/contracts/<int:cid>/edit', methods=['GET', 'POST'])
@login_required
def contract_edit(cid):
    if not can_edit('contract', cid, current_user):
        abort(403)
    db = get_db()
    c = db.execute('SELECT * FROM contracts WHERE id=?', (cid,)).fetchone()
    if not c:
        abort(404)
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        if not title:
            flash('Tên hợp đồng không được trống', 'danger')
            return redirect(url_for('contract_edit', cid=cid))
        files_state = {
            'contract_file': c['contract_file'],
            'handover_file': c['handover_file'],
            'appendix_file': c['appendix_file'],
            'invoice_file': c['invoice_file'],
            'other_file': c['other_file'],
        }
        for fld in files_state:
            f = request.files.get(fld)
            if f and f.filename:
                s, _ = save_upload(f, CONTRACT_DIR)
                if s:
                    files_state[fld] = s
        try:
            pct = int(request.form.get('progress_percent') or 0)
            pct = max(0, min(100, pct))
        except ValueError:
            pct = 0
        rcv_ids, rcv_names, rcv_emails = receivers_from_form(request.form)
        db.execute('''UPDATE contracts SET code=?, title=?, partner=?, supplier_tax=?, contact_person=?,
                     contact_phone=?, receiver=?, receiver_email=?, receiver_ids=?,
                     total_value=?, paid_amount=?, contract_date=?, due_date=?,
                     progress=?, progress_percent=?, notes=?, contract_file=?, handover_file=?,
                     appendix_file=?, invoice_file=?, other_file=?, updated_at=? WHERE id=?''',
                   (request.form.get('code', '').strip(),
                    title,
                    request.form.get('partner', '').strip(),
                    request.form.get('supplier_tax', '').strip(),
                    request.form.get('contact_person', '').strip(),
                    request.form.get('contact_phone', '').strip(),
                    rcv_names, rcv_emails,
                    ','.join(str(i) for i in rcv_ids) or None,
                    _parse_money(request.form.get('total_value')),
                    _parse_money(request.form.get('paid_amount')),
                    request.form.get('contract_date') or None,
                    request.form.get('due_date') or None,
                    request.form.get('progress', '').strip(),
                    pct,
                    request.form.get('notes', '').strip(),
                    files_state['contract_file'], files_state['handover_file'],
                    files_state['appendix_file'], files_state['invoice_file'], files_state['other_file'],
                    datetime.now().isoformat(timespec='seconds'), cid))
        db.commit()
        flash('Đã cập nhật hợp đồng', 'success')
        return redirect(url_for('contract_view', cid=cid))
    sel = [int(i) for i in (c['receiver_ids'] or '').split(',') if i.strip().isdigit()]
    return render_template('contract_form.html', contract=c,
                           users=_receiver_user_list(), selected_receiver_ids=sel)


@app.route('/contracts/<int:cid>/delete', methods=['POST'])
@login_required
def contract_delete(cid):
    if not can_edit('contract', cid, current_user):
        abort(403)
    db = get_db()
    c = db.execute('''SELECT owner_id, contract_file, handover_file, appendix_file,
                             invoice_file, other_file FROM contracts WHERE id=?''',
                   (cid,)).fetchone()
    if not c:
        abort(404)
    if not current_user.full_access and c['owner_id'] != current_user.id:
        flash('Chỉ chủ hợp đồng hoặc Admin mới được xoá', 'danger')
        return redirect(url_for('contract_view', cid=cid))
    db.execute('DELETE FROM record_permissions WHERE record_type=? AND record_id=?', ('contract', cid))
    db.execute('DELETE FROM contracts WHERE id=?', (cid,))
    db.commit()
    for f in (c['contract_file'], c['handover_file'], c['appendix_file'],
              c['invoice_file'], c['other_file']):
        if f:
            try:
                os.remove(os.path.join(CONTRACT_DIR, f))
            except OSError:
                pass
    flash('Đã xoá hợp đồng', 'success')
    return redirect(url_for('contracts_list'))


@app.route('/contracts/<int:cid>/file/<which>')
@login_required
def contract_file(cid, which):
    if which not in ('contract_file', 'handover_file', 'appendix_file', 'invoice_file', 'other_file'):
        abort(400)
    if not can_view('contract', cid, current_user):
        abort(403)
    c = get_db().execute(f'SELECT {which} AS f FROM contracts WHERE id=?', (cid,)).fetchone()
    if not c or not c['f']:
        abort(404)
    return send_from_directory(CONTRACT_DIR, c['f'], as_attachment=True)


# ---------- Sharing ----------
@app.route('/share/<rtype>/<int:rid>', methods=['POST'])
@login_required
def share_record(rtype, rid):
    if rtype not in ('contract', 'order', 'media'):
        abort(400)
    perm = get_perm(rtype, rid, current_user)
    if perm != 'owner':
        flash('Chỉ chủ hồ sơ hoặc Admin mới được phân quyền', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    user_id = request.form.get('user_id', type=int)
    perm_value = request.form.get('perm', 'view')
    if perm_value not in ('view', 'edit'):
        abort(400)
    if not user_id:
        flash('Chưa chọn user', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    db = get_db()
    db.execute('''INSERT INTO record_permissions(record_type, record_id, user_id, perm)
                 VALUES (?,?,?,?)
                 ON CONFLICT(record_type, record_id, user_id) DO UPDATE SET perm=excluded.perm''',
               (rtype, rid, user_id, perm_value))
    db.commit()
    flash('Đã cập nhật quyền', 'success')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/share/<rtype>/<int:rid>/revoke/<int:user_id>', methods=['POST'])
@login_required
def revoke_share(rtype, rid, user_id):
    if rtype not in ('contract', 'order', 'media'):
        abort(400)
    if get_perm(rtype, rid, current_user) != 'owner':
        abort(403)
    db = get_db()
    db.execute('DELETE FROM record_permissions WHERE record_type=? AND record_id=? AND user_id=?',
               (rtype, rid, user_id))
    db.commit()
    flash('Đã thu hồi quyền', 'success')
    return redirect(request.referrer or url_for('dashboard'))


# ---------- Attachments ----------
@app.route('/attach/<rtype>/<int:rid>/<category>', methods=['POST'])
@login_required
def attach_upload(rtype, rid, category):
    if rtype != 'contract':
        abort(400)
    if category not in DOC_CATEGORY_KEYS:
        abort(400)
    if not can_edit(rtype, rid, current_user):
        abort(403)
    files = request.files.getlist('files')
    db = get_db()
    saved = 0
    for f in files:
        if not f or not f.filename:
            continue
        stored, orig = save_upload(f, ATTACH_DIR)
        if stored:
            db.execute('''INSERT INTO attachments(record_type, record_id, category, stored_name, original_name, created_at)
                          VALUES (?,?,?,?,?,?)''',
                       (rtype, rid, category, stored, orig, datetime.now().isoformat(timespec='seconds')))
            saved += 1
    db.commit()
    flash(f'Đã tải lên {saved} file', 'success' if saved else 'warning')
    return redirect(url_for('contract_view', cid=rid) + '#files')


@app.route('/attach/<int:aid>/download')
@login_required
def attach_download(aid):
    a = get_db().execute('SELECT * FROM attachments WHERE id=?', (aid,)).fetchone()
    if not a:
        abort(404)
    if not can_view(a['record_type'], a['record_id'], current_user):
        abort(403)
    return send_from_directory(ATTACH_DIR, a['stored_name'],
                               as_attachment=True, download_name=a['original_name'] or a['stored_name'])


@app.route('/attach/<int:aid>/delete', methods=['POST'])
@login_required
def attach_delete(aid):
    db = get_db()
    a = db.execute('SELECT * FROM attachments WHERE id=?', (aid,)).fetchone()
    if not a:
        abort(404)
    if not can_edit(a['record_type'], a['record_id'], current_user):
        abort(403)
    db.execute('DELETE FROM attachments WHERE id=?', (aid,))
    db.commit()
    try:
        os.remove(os.path.join(ATTACH_DIR, a['stored_name']))
    except OSError:
        pass
    flash('Đã xoá file', 'success')
    return redirect(url_for('contract_view', cid=a['record_id']) + '#files')


# ---------- Comments ----------
@app.route('/comment/<rtype>/<int:rid>', methods=['POST'])
@login_required
def add_comment(rtype, rid):
    if rtype not in ('contract', 'order', 'media'):
        abort(400)
    if not can_view(rtype, rid, current_user):
        abort(403)
    if not current_user.has_cap('comment'):
        flash('Bạn không có quyền nhập nhận xét', 'danger')
        return redirect(_record_back_url(rtype, rid) + '#comments')
    content = request.form.get('content', '').strip()
    if not content:
        flash('Nhận xét không được để trống', 'danger')
    else:
        db = get_db()
        db.execute('''INSERT INTO comments(record_type, record_id, user_id, content, created_at)
                     VALUES (?,?,?,?,?)''',
                   (rtype, rid, current_user.id, content,
                    datetime.now().isoformat(timespec='seconds')))
        db.commit()
        flash('Đã thêm nhận xét', 'success')
    return redirect(_record_back_url(rtype, rid) + '#comments')


def _record_back_url(rtype, rid):
    if rtype == 'contract':
        return url_for('contract_view', cid=rid)
    elif rtype == 'order':
        return url_for('order_view', oid=rid)
    return url_for('media_view', mid=rid)


@app.route('/comment/<int:cmt_id>/delete', methods=['POST'])
@login_required
def delete_comment(cmt_id):
    db = get_db()
    c = db.execute('SELECT * FROM comments WHERE id=?', (cmt_id,)).fetchone()
    if not c:
        abort(404)
    if not current_user.is_admin and c['user_id'] != current_user.id:
        flash('Chỉ tác giả nhận xét hoặc Admin mới được xoá', 'danger')
    else:
        db.execute('DELETE FROM comments WHERE id=?', (cmt_id,))
        db.commit()
        flash('Đã xoá nhận xét', 'success')
    return redirect(_record_back_url(c['record_type'], c['record_id']) + '#comments')


def get_comments(record_type, record_id):
    return get_db().execute('''
        SELECT c.*, u.username, u.full_name, u.organization
        FROM comments c JOIN users u ON c.user_id = u.id
        WHERE c.record_type=? AND c.record_id=?
        ORDER BY c.created_at ASC
    ''', (record_type, record_id)).fetchall()


# ============================================================
# ---------- TRUYỀN THÔNG (Media) ----------
# ============================================================
@app.route('/media')
@login_required
def media_list():
    db = get_db()
    ids = list_accessible_ids('media', current_user)
    status_filter = request.args.get('status', '').strip()
    where = ''
    params = []
    if status_filter and status_filter in MEDIA_STATUS:
        where = ' WHERE m.status=?'
        params.append(status_filter)
    if ids is None:
        base = ('''SELECT m.*, u.username owner_user, u.full_name owner_name, u.organization owner_org
                   FROM media_posts m JOIN users u ON m.owner_id=u.id''' + where +
                ' ORDER BY m.updated_at DESC')
        rows = db.execute(base, params).fetchall()
    elif ids:
        placeholder = ','.join('?' * len(ids))
        if where:
            full_where = where + f' AND m.id IN ({placeholder})'
        else:
            full_where = f' WHERE m.id IN ({placeholder})'
        q = ('''SELECT m.*, u.username owner_user, u.full_name owner_name, u.organization owner_org
                FROM media_posts m JOIN users u ON m.owner_id=u.id''' + full_where +
             ' ORDER BY m.updated_at DESC')
        rows = db.execute(q, params + list(ids)).fetchall()
    else:
        rows = []
    return render_template('media_list.html', posts=rows, current_status=status_filter)


@app.route('/media/new', methods=['GET', 'POST'])
@login_required
@require_cap('create_media')
def media_new():
    if request.method == 'POST':
        return _save_media(None)
    return render_template('media_form.html', post=None, files=[])


def _save_media(post_id):
    db = get_db()
    is_new = post_id is None
    title = request.form.get('title', '').strip()
    if not title:
        flash('Tiêu đề không được trống', 'danger')
        return redirect(url_for('media_new') if not post_id else url_for('media_edit', mid=post_id))
    content = request.form.get('content', '').strip()
    channel = request.form.get('channel', '').strip()
    planned_date = request.form.get('planned_date') or None
    link_val = request.form.get('link', '').strip()
    notes = request.form.get('notes', '').strip()
    now = datetime.now().isoformat(timespec='seconds')

    if is_new:
        cur = db.execute('''INSERT INTO media_posts
            (title, content, channel, planned_date, link, notes, status, owner_id, created_at, updated_at)
            VALUES (?,?,?,?,?,?, 'draft', ?,?,?)''',
            (title, content, channel, planned_date, link_val, notes, current_user.id, now, now))
        post_id = cur.lastrowid
    else:
        db.execute('''UPDATE media_posts SET title=?, content=?, channel=?, planned_date=?,
                      link=?, notes=?, updated_at=? WHERE id=?''',
                   (title, content, channel, planned_date, link_val, notes, now, post_id))

    # Upload kèm file mới (giữ các file cũ)
    for f in request.files.getlist('files'):
        if not f or not f.filename:
            continue
        stored, orig = save_upload(f, MEDIA_DIR)
        if stored:
            db.execute('''INSERT INTO media_files(post_id, stored_name, original_name, uploaded_by, created_at)
                          VALUES (?,?,?,?,?)''',
                       (post_id, stored, orig, current_user.id, now))
    db.commit()
    flash('Đã lưu bài truyền thông', 'success')
    return redirect(url_for('media_view', mid=post_id))


@app.route('/media/<int:mid>')
@login_required
def media_view(mid):
    if not can_view('media', mid, current_user):
        abort(403)
    db = get_db()
    m = db.execute('''SELECT m.*, u.username owner_user, u.full_name owner_name, u.organization owner_org,
                             cf.full_name confirmer_name, cf.organization confirmer_org
                      FROM media_posts m JOIN users u ON m.owner_id=u.id
                      LEFT JOIN users cf ON m.confirmed_by=cf.id
                      WHERE m.id=?''', (mid,)).fetchone()
    if not m:
        abort(404)
    files = db.execute('SELECT * FROM media_files WHERE post_id=? ORDER BY id', (mid,)).fetchall()
    comments = get_comments('media', mid)
    return render_template('media_view.html', m=m, files=files, comments=comments,
                           editable=can_edit('media', mid, current_user))


@app.route('/media/<int:mid>/edit', methods=['GET', 'POST'])
@login_required
def media_edit(mid):
    if not can_edit('media', mid, current_user):
        abort(403)
    db = get_db()
    m = db.execute('SELECT * FROM media_posts WHERE id=?', (mid,)).fetchone()
    if not m:
        abort(404)
    # Chỉ sửa được khi draft/revision (hoặc admin)
    if m['status'] not in ('draft', 'revision') and not current_user.is_admin:
        flash('Chỉ sửa được bài ở trạng thái Nháp hoặc KBC yêu cầu sửa', 'warning')
        return redirect(url_for('media_view', mid=mid))
    if request.method == 'POST':
        return _save_media(mid)
    files = db.execute('SELECT * FROM media_files WHERE post_id=? ORDER BY id', (mid,)).fetchall()
    return render_template('media_form.html', post=m, files=files)


@app.route('/media/<int:mid>/delete', methods=['POST'])
@login_required
def media_delete(mid):
    if not can_edit('media', mid, current_user):
        abort(403)
    db = get_db()
    m = db.execute('SELECT owner_id FROM media_posts WHERE id=?', (mid,)).fetchone()
    if not m:
        abort(404)
    if not current_user.full_access and m['owner_id'] != current_user.id:
        flash('Chỉ chủ bài hoặc Admin mới được xoá', 'danger')
        return redirect(url_for('media_view', mid=mid))
    files = db.execute('SELECT stored_name FROM media_files WHERE post_id=?', (mid,)).fetchall()
    db.execute('DELETE FROM media_posts WHERE id=?', (mid,))
    db.commit()
    for f in files:
        try:
            os.remove(os.path.join(MEDIA_DIR, f['stored_name']))
        except OSError:
            pass
    flash('Đã xoá bài truyền thông', 'success')
    return redirect(url_for('media_list'))


@app.route('/media/file/<int:fid>/download')
@login_required
def media_file_download(fid):
    db = get_db()
    f = db.execute('SELECT * FROM media_files WHERE id=?', (fid,)).fetchone()
    if not f:
        abort(404)
    if not can_view('media', f['post_id'], current_user):
        abort(403)
    return send_from_directory(MEDIA_DIR, f['stored_name'],
                               as_attachment=True,
                               download_name=f['original_name'] or f['stored_name'])


@app.route('/media/file/<int:fid>/delete', methods=['POST'])
@login_required
def media_file_delete(fid):
    db = get_db()
    f = db.execute('SELECT * FROM media_files WHERE id=?', (fid,)).fetchone()
    if not f:
        abort(404)
    if not can_edit('media', f['post_id'], current_user):
        abort(403)
    db.execute('DELETE FROM media_files WHERE id=?', (fid,))
    db.commit()
    try:
        os.remove(os.path.join(MEDIA_DIR, f['stored_name']))
    except OSError:
        pass
    flash('Đã xoá file', 'success')
    return redirect(url_for('media_view', mid=f['post_id']))


@app.route('/media/<int:mid>/submit', methods=['POST'])
@login_required
def media_submit(mid):
    """HP89 gửi bài đi để KBC duyệt."""
    db = get_db()
    m = db.execute('SELECT * FROM media_posts WHERE id=?', (mid,)).fetchone()
    if not m:
        abort(404)
    if not (current_user.is_admin or m['owner_id'] == current_user.id):
        abort(403)
    if m['status'] not in ('draft', 'revision'):
        flash('Bài đang ở trạng thái không thể gửi duyệt', 'warning')
        return redirect(url_for('media_view', mid=mid))
    now = datetime.now().isoformat(timespec='seconds')
    db.execute("UPDATE media_posts SET status='pending', submitted_at=?, updated_at=? WHERE id=?",
               (now, now, mid))
    link = url_for('media_view', mid=mid)
    notify_cap_users('approve_media',
                     f'[TRUYỀN THÔNG] HP89 gửi bài "{m["title"]}" — cần KBC xác nhận',
                     link, exclude_ids=[current_user.id], organization='KBC')
    db.commit()
    flash('Đã gửi nội dung cho KBC duyệt', 'success')
    return redirect(url_for('media_view', mid=mid))


@app.route('/media/<int:mid>/approve', methods=['POST'])
@login_required
@require_cap('approve_media')
def media_approve(mid):
    """KBC xác nhận thống nhất nội dung."""
    db = get_db()
    m = db.execute('SELECT * FROM media_posts WHERE id=?', (mid,)).fetchone()
    if not m:
        abort(404)
    now = datetime.now().isoformat(timespec='seconds')
    db.execute('''UPDATE media_posts SET status='approved', confirmed_by=?, confirmed_at=?,
                  revision_note=NULL, updated_at=? WHERE id=?''',
               (current_user.id, now, now, mid))
    link = url_for('media_view', mid=mid)
    notify_user(m['owner_id'],
                f'KBC đã THỐNG NHẤT nội dung truyền thông "{m["title"]}"', link)
    db.commit()
    flash('Đã xác nhận thống nhất nội dung', 'success')
    return redirect(url_for('media_view', mid=mid))


@app.route('/media/<int:mid>/request-revision', methods=['POST'])
@login_required
@require_cap('approve_media')
def media_request_revision(mid):
    """KBC yêu cầu HP89 chỉnh sửa."""
    db = get_db()
    m = db.execute('SELECT * FROM media_posts WHERE id=?', (mid,)).fetchone()
    if not m:
        abort(404)
    note = request.form.get('revision_note', '').strip()
    now = datetime.now().isoformat(timespec='seconds')
    db.execute('''UPDATE media_posts SET status='revision', revision_note=?, updated_at=? WHERE id=?''',
               (note, now, mid))
    link = url_for('media_view', mid=mid)
    msg = f'KBC yêu cầu CHỈNH SỬA bài "{m["title"]}"'
    if note:
        msg += f' — Lý do: {note}'
    notify_user(m['owner_id'], msg, link)
    db.commit()
    flash('Đã gửi yêu cầu chỉnh sửa về HP89', 'warning')
    return redirect(url_for('media_view', mid=mid))


@app.route('/media/<int:mid>/publish', methods=['POST'])
@login_required
def media_publish(mid):
    """Đánh dấu bài đã đăng (sau khi KBC đã duyệt)."""
    db = get_db()
    m = db.execute('SELECT * FROM media_posts WHERE id=?', (mid,)).fetchone()
    if not m:
        abort(404)
    if not (current_user.is_admin or m['owner_id'] == current_user.id):
        abort(403)
    if m['status'] != 'approved':
        flash('Chỉ bài đã được KBC duyệt mới đánh dấu được Đã đăng', 'warning')
        return redirect(url_for('media_view', mid=mid))
    link_val = request.form.get('link', '').strip()
    now = datetime.now().isoformat(timespec='seconds')
    db.execute("UPDATE media_posts SET status='published', link=?, updated_at=? WHERE id=?",
               (link_val, now, mid))
    db.commit()
    flash('Đã đánh dấu bài Đã đăng', 'success')
    return redirect(url_for('media_view', mid=mid))


# ============================================================
# ---------- PRODUCTS (Danh mục sản phẩm) ----------
# ============================================================
@app.route('/products')
@login_required
@require_cap('manage_products')
def products_list():
    rows = get_db().execute('SELECT * FROM products ORDER BY active DESC, name ASC').fetchall()
    return render_template('products_list.html', products=rows)


@app.route('/products/new', methods=['GET', 'POST'])
@login_required
@require_cap('manage_products')
def product_new():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Tên sản phẩm không được trống', 'danger')
            return redirect(url_for('product_new'))
        try:
            price = float(request.form.get('default_price') or 0)
        except ValueError:
            price = 0
        db = get_db()
        db.execute('''INSERT INTO products(name, unit, packaging, default_price, active, created_at)
                     VALUES (?,?,?,?,1,?)''',
                   (name, request.form.get('unit', '').strip(),
                    request.form.get('packaging', '').strip(),
                    price, datetime.now().isoformat(timespec='seconds')))
        db.commit()
        flash(f'Đã thêm sản phẩm "{name}"', 'success')
        return redirect(url_for('products_list'))
    return render_template('product_form.html', product=None)


@app.route('/products/<int:prod_id>/edit', methods=['GET', 'POST'])
@login_required
@require_cap('manage_products')
def product_edit(prod_id):
    db = get_db()
    p = db.execute('SELECT * FROM products WHERE id=?', (prod_id,)).fetchone()
    if not p:
        abort(404)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Tên sản phẩm không được trống', 'danger')
            return redirect(url_for('product_edit', prod_id=prod_id))
        try:
            price = float(request.form.get('default_price') or 0)
        except ValueError:
            price = 0
        active = 1 if request.form.get('active') else 0
        db.execute('''UPDATE products SET name=?, unit=?, packaging=?, default_price=?, active=?
                     WHERE id=?''',
                   (name, request.form.get('unit', '').strip(),
                    request.form.get('packaging', '').strip(),
                    price, active, prod_id))
        db.commit()
        flash('Đã cập nhật sản phẩm', 'success')
        return redirect(url_for('products_list'))
    return render_template('product_form.html', product=p)


@app.route('/products/<int:prod_id>/delete', methods=['POST'])
@login_required
@require_cap('manage_products')
def product_delete(prod_id):
    db = get_db()
    db.execute('DELETE FROM products WHERE id=?', (prod_id,))
    db.commit()
    flash('Đã xoá sản phẩm', 'success')
    return redirect(url_for('products_list'))


@app.route('/api/products')
@login_required
def api_products():
    rows = get_db().execute('SELECT id, name, unit, packaging, default_price FROM products WHERE active=1 ORDER BY name').fetchall()
    return jsonify([dict(r) for r in rows])


# ============================================================
# ---------- INVOICE ENTITIES (Thông tin đơn vị xuất HĐ) ----------
# ============================================================
@app.route('/api/invoice-entities', methods=['GET', 'POST'])
@login_required
def api_invoice_entities():
    db = get_db()
    if request.method == 'POST':
        data = request.get_json() or {}
        company_name = (data.get('company_name') or '').strip()
        if not company_name:
            return jsonify({'error': 'Tên đơn vị không được trống'}), 400
        now = datetime.now().isoformat(timespec='seconds')
        cur = db.execute(
            'INSERT INTO invoice_entities(company_name, tax_code, address, email, phone, created_by, created_at) VALUES (?,?,?,?,?,?,?)',
            (company_name, (data.get('tax_code') or '').strip(),
             (data.get('address') or '').strip(),
             (data.get('email') or '').strip(),
             (data.get('phone') or '').strip(),
             current_user.id, now))
        db.commit()
        return jsonify({'id': cur.lastrowid, 'company_name': company_name,
                        'tax_code': data.get('tax_code', ''),
                        'address': data.get('address', ''),
                        'email': data.get('email', ''),
                        'phone': data.get('phone', '')})
    rows = db.execute('SELECT * FROM invoice_entities ORDER BY company_name').fetchall()
    return jsonify([dict(r) for r in rows])


# ============================================================
# ---------- NOTIFICATIONS ----------
# ============================================================
@app.route('/notifications')
@login_required
def notifications_list():
    rows = get_db().execute(
        'SELECT * FROM notifications WHERE user_id=? ORDER BY is_read ASC, created_at DESC LIMIT 100',
        (current_user.id,)).fetchall()
    return render_template('notifications.html', notifications=rows)


@app.route('/notifications/count')
@login_required
def notifications_count():
    """API trả số notification chưa đọc — dùng cho client poll phát âm khi có tin mới."""
    row = get_db().execute(
        'SELECT COUNT(*) c FROM notifications WHERE user_id=? AND is_read=0',
        (current_user.id,)).fetchone()
    return jsonify(unread=(row['c'] if row else 0))


@app.route('/notifications/<int:nid>/open')
@login_required
def notification_open(nid):
    db = get_db()
    n = db.execute('SELECT * FROM notifications WHERE id=? AND user_id=?',
                   (nid, current_user.id)).fetchone()
    if not n:
        abort(404)
    db.execute('UPDATE notifications SET is_read=1 WHERE id=?', (nid,))
    db.commit()
    return redirect(n['link'] or url_for('notifications_list'))


@app.route('/notifications/read-all', methods=['POST'])
@login_required
def notifications_read_all():
    db = get_db()
    db.execute('UPDATE notifications SET is_read=1 WHERE user_id=?', (current_user.id,))
    db.commit()
    flash('Đã đánh dấu tất cả thông báo là đã đọc', 'success')
    return redirect(url_for('notifications_list'))


# ---------- Web Push ----------
@app.route('/push/subscribe', methods=['POST'])
@login_required
def push_subscribe():
    data = request.get_json(silent=True) or {}
    sub = data.get('subscription')
    if not sub or not sub.get('endpoint'):
        return jsonify(ok=False, error='Thiếu thông tin đăng ký'), 400
    endpoint = sub['endpoint']
    sub_json = json.dumps(sub)
    now = datetime.now().isoformat(timespec='seconds')
    db = get_db()
    existing = db.execute('SELECT id FROM push_subscriptions WHERE endpoint=?', (endpoint,)).fetchone()
    if existing:
        db.execute('UPDATE push_subscriptions SET user_id=?, subscription=?, created_at=? WHERE endpoint=?',
                   (current_user.id, sub_json, now, endpoint))
    else:
        db.execute('INSERT INTO push_subscriptions(user_id, endpoint, subscription, created_at) VALUES (?,?,?,?)',
                   (current_user.id, endpoint, sub_json, now))
    db.commit()
    return jsonify(ok=True)


@app.route('/push/unsubscribe', methods=['POST'])
@login_required
def push_unsubscribe():
    data = request.get_json(silent=True) or {}
    endpoint = data.get('endpoint')
    if endpoint:
        db = get_db()
        db.execute('DELETE FROM push_subscriptions WHERE endpoint=?', (endpoint,))
        db.commit()
    return jsonify(ok=True)


@app.route('/push/test', methods=['POST'])
@login_required
def push_test():
    send_web_push_to_user(current_user.id, 'KBC-HP89',
                          'Thông báo thử - nếu thấy dòng này thì đã hoạt động!',
                          url_for('notifications_list'))
    get_db().commit()
    return jsonify(ok=True)


# ============================================================
# ---------- REPORTS ----------
# ============================================================
@app.route('/reports')
@login_required
def reports():
    return render_template('reports.html')


@app.route('/reports/orders.xlsx')
@login_required
def report_orders():
    db = get_db()
    ids = list_accessible_ids('order', current_user)
    if ids is None:
        rows = db.execute('''SELECT o.*, u.username owner_user, u.organization owner_org FROM orders o
                             JOIN users u ON o.owner_id=u.id ORDER BY o.order_date DESC''').fetchall()
    elif ids:
        q = f'''SELECT o.*, u.username owner_user, u.organization owner_org FROM orders o
                JOIN users u ON o.owner_id=u.id WHERE o.id IN ({','.join('?' * len(ids))})
                ORDER BY o.order_date DESC'''
        rows = db.execute(q, list(ids)).fetchall()
    else:
        rows = []
    wb = Workbook()
    ws = wb.active
    ws.title = 'Đơn hàng KBC-HP89'
    ws.append(['ID', 'Mã ĐH', 'Ngày', 'Nơi nhận', 'SĐT', 'Địa chỉ',
               'Đơn vị xuất HĐ', 'MST', 'Hình thức TT',
               'Trước VAT', 'VAT %', 'VAT', 'Tổng cộng', 'Đã thanh toán',
               'Trạng thái', 'Người tạo', 'Đơn vị'])
    for r in rows:
        ws.append([r['id'], r['code'], r['order_date'], r['customer_name'],
                   r['customer_phone'], r['customer_address'],
                   r['invoice_company'], r['invoice_tax_code'], r['payment_method'],
                   r['subtotal'], r['vat_percent'], r['vat_amount'], r['grand_total'],
                   r['paid_amount'],
                   ORDER_WF_LABEL.get(r['workflow_status'], r['workflow_status']),
                   r['owner_user'], r['owner_org']])
    ws2 = wb.create_sheet('Chi tiết')
    ws2.append(['Mã ĐH', 'Nơi nhận', 'STT', 'Sản phẩm', 'ĐVT', 'Quy cách',
                'SL', 'Đơn giá', 'Thành tiền', 'CK %', 'Thành tiền sau CK', 'Ghi chú'])
    for r in rows:
        items = db.execute('SELECT * FROM order_items WHERE order_id=? ORDER BY sort_order, id', (r['id'],)).fetchall()
        for idx, it in enumerate(items, 1):
            ws2.append([r['code'], r['customer_name'], idx, it['product_name'],
                        it['unit'], it['packaging'], it['qty'], it['unit_price'],
                        it['amount'], it['discount'], it['amount_after'], it['note']])
    for col_idx in range(1, 18):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = 16
    for col_idx in range(1, 13):
        ws2.column_dimensions[chr(64 + col_idx)].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f'bao_cao_don_hang_{date.today().isoformat()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/reports/contracts.xlsx')
@login_required
def report_contracts():
    db = get_db()
    ids = list_accessible_ids('contract', current_user)
    if ids is None:
        rows = db.execute('''SELECT c.*, u.username owner_user, u.organization owner_org FROM contracts c
                             JOIN users u ON c.owner_id=u.id ORDER BY c.due_date ASC''').fetchall()
    elif ids:
        q = f'''SELECT c.*, u.username owner_user, u.organization owner_org FROM contracts c
                JOIN users u ON c.owner_id=u.id WHERE c.id IN ({','.join('?' * len(ids))})
                ORDER BY c.due_date ASC'''
        rows = db.execute(q, list(ids)).fetchall()
    else:
        rows = []
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hợp đồng'
    ws.append(['ID', 'Mã HĐ', 'Tên hợp đồng', 'Đối tác', 'Ngày HĐ', 'Hạn HĐ',
               'Giá trị', 'Đã TT', 'Tiến độ', '%', 'Trạng thái', 'Chủ HĐ', 'Đơn vị', 'Ghi chú'])
    for r in rows:
        st, days = contract_status(r['due_date'])
        st_text = {'overdue': f'TRỄ HẠN ({-days} ngày)' if days is not None else 'TRỄ HẠN',
                   'soon': f'Sắp đến hạn ({days} ngày)',
                   'ok': 'Còn hạn',
                   'none': ''}[st]
        ws.append([r['id'], r['code'], r['title'], r['partner'],
                   r['contract_date'], r['due_date'],
                   r['total_value'], r['paid_amount'],
                   r['progress'], r['progress_percent'], st_text,
                   r['owner_user'], r['owner_org'], r['notes']])
    for col in 'ABCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f'bao_cao_hop_dong_{date.today().isoformat()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/reports/media.xlsx')
@login_required
def report_media():
    db = get_db()
    ids = list_accessible_ids('media', current_user)
    if ids is None:
        rows = db.execute('''SELECT m.*, u.username owner_user, u.organization owner_org FROM media_posts m
                             JOIN users u ON m.owner_id=u.id ORDER BY m.updated_at DESC''').fetchall()
    elif ids:
        q = f'''SELECT m.*, u.username owner_user, u.organization owner_org FROM media_posts m
                JOIN users u ON m.owner_id=u.id WHERE m.id IN ({','.join('?' * len(ids))})
                ORDER BY m.updated_at DESC'''
        rows = db.execute(q, list(ids)).fetchall()
    else:
        rows = []
    wb = Workbook()
    ws = wb.active
    ws.title = 'Truyền thông'
    ws.append(['ID', 'Tiêu đề', 'Kênh', 'Ngày dự kiến', 'Link', 'Trạng thái',
               'Người đăng', 'Đơn vị', 'Ngày tạo', 'Cập nhật'])
    for r in rows:
        ws.append([r['id'], r['title'], r['channel'], r['planned_date'], r['link'],
                   MEDIA_STATUS_LABEL.get(r['status'], r['status']),
                   r['owner_user'], r['owner_org'], r['created_at'], r['updated_at']])
    for col in 'ABCDEFGHIJ':
        ws.column_dimensions[col].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f'bao_cao_truyen_thong_{date.today().isoformat()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# ---------- LEGAL (Pháp Lý) ----------
# ============================================================
def _legal_tree(db, parent_id=None, depth=0):
    rows = db.execute(
        '''SELECT n.*, u.full_name AS creator_name, a.full_name AS approver_name
           FROM legal_nodes n
           LEFT JOIN users u ON u.id = n.created_by
           LEFT JOIN users a ON a.id = n.approved_by
           WHERE n.parent_id IS ?
           ORDER BY n.created_at''',
        (parent_id,)
    ).fetchall()
    result = []
    for r in rows:
        node = dict(r)
        node['depth'] = depth
        node['children'] = _legal_tree(db, r['id'], depth + 1)
        result.append(node)
    return result


def _legal_indent_options(nodes, prefix=''):
    result = []
    for n in nodes:
        result.append((n['id'], prefix + n['title']))
        result += _legal_indent_options(n['children'], prefix + '— ')
    return result


@app.route('/legal')
@login_required
def legal_list():
    db = get_db()
    tree = _legal_tree(db)
    return render_template('legal_list.html', tree=tree)


@app.route('/legal/<int:nid>')
@login_required
def legal_view(nid):
    db = get_db()
    node = db.execute(
        '''SELECT n.*, u.full_name AS creator_name, a.full_name AS approver_name
           FROM legal_nodes n
           LEFT JOIN users u ON u.id = n.created_by
           LEFT JOIN users a ON a.id = n.approved_by
           WHERE n.id=?''', (nid,)).fetchone()
    if not node:
        abort(404)
    children = db.execute(
        '''SELECT id, title, description, approval_status, stored_name
           FROM legal_nodes WHERE parent_id=? ORDER BY created_at''', (nid,)).fetchall()
    crumbs = []
    cur = node
    while cur is not None:
        crumbs.append({'id': cur['id'], 'title': cur['title']})
        parent_id = cur['parent_id']
        cur = db.execute('SELECT id, title, parent_id FROM legal_nodes WHERE id=?',
                         (parent_id,)).fetchone() if parent_id else None
    crumbs.reverse()
    return render_template('legal_view.html', node=node, children=children, crumbs=crumbs)


@app.route('/legal/<int:nid>/preview')
@login_required
def legal_preview(nid):
    db = get_db()
    node = db.execute('SELECT stored_name, original_name FROM legal_nodes WHERE id=?', (nid,)).fetchone()
    if not node or not node['stored_name']:
        abort(404)
    return send_from_directory(LEGAL_DIR, node['stored_name'], as_attachment=False,
                               download_name=node['original_name'] or node['stored_name'])


@app.route('/legal/new', methods=['GET', 'POST'])
@login_required
def legal_new():
    db = get_db()
    tree = _legal_tree(db)
    options = _legal_indent_options(tree)

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        parent_id = request.form.get('parent_id') or None
        description = request.form.get('description', '').strip()
        if not title:
            flash('Vui lòng nhập tiêu đề.', 'danger')
            return render_template('legal_form.html', options=options, action='new',
                                   form=request.form)

        if parent_id:
            parent_id = int(parent_id)
            if not db.execute('SELECT 1 FROM legal_nodes WHERE id=?', (parent_id,)).fetchone():
                parent_id = None

        stored_name = None
        original_name = None
        f = request.files.get('file')
        if f and f.filename:
            ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
            if ext not in ALLOWED_EXT:
                flash(f'Định dạng file .{ext} không được phép.', 'danger')
                return render_template('legal_form.html', options=options, action='new',
                                       form=request.form)
            original_name = secure_filename(f.filename)
            stored_name = f'{datetime.now().strftime("%Y%m%d%H%M%S%f")}_{original_name}'
            f.save(os.path.join(LEGAL_DIR, stored_name))

        now = datetime.now().isoformat(timespec='seconds')
        db.execute(
            '''INSERT INTO legal_nodes(title, parent_id, description, stored_name, original_name,
               approval_status, created_by, created_at, updated_at)
               VALUES (?,?,?,?,?,'pending',?,?,?)''',
            (title, parent_id, description, stored_name, original_name,
             current_user.id, now, now)
        )
        db.commit()
        # Báo cho người có quyền duyệt
        notify_cap_users('approve_legal', f'Giấy tờ pháp lý mới «{title}» cần duyệt',
                         url_for('legal_list'), exclude_ids=[current_user.id])
        db.commit()
        flash('Đã tạo giấy tờ pháp lý. Đang chờ phê duyệt.', 'success')
        return redirect(url_for('legal_list'))

    return render_template('legal_form.html', options=options, action='new', form={})


@app.route('/legal/<int:nid>/edit', methods=['GET', 'POST'])
@login_required
def legal_edit(nid):
    db = get_db()
    node = db.execute('SELECT * FROM legal_nodes WHERE id=?', (nid,)).fetchone()
    if not node:
        abort(404)
    tree = _legal_tree(db)
    options = _legal_indent_options(tree)
    options = [(oid, lbl) for oid, lbl in options if oid != nid]

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        parent_id = request.form.get('parent_id') or None
        description = request.form.get('description', '').strip()
        if not title:
            flash('Vui lòng nhập tiêu đề.', 'danger')
            return render_template('legal_form.html', options=options, action='edit',
                                   node=node, form=request.form)
        if parent_id:
            parent_id = int(parent_id)
            if parent_id == nid:
                parent_id = None
        stored_name = node['stored_name']
        original_name = node['original_name']
        f = request.files.get('file')
        if f and f.filename:
            ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
            if ext not in ALLOWED_EXT:
                flash(f'Định dạng file .{ext} không được phép.', 'danger')
                return render_template('legal_form.html', options=options, action='edit',
                                       node=node, form=request.form)
            if stored_name:
                old_path = os.path.join(LEGAL_DIR, stored_name)
                if os.path.exists(old_path):
                    os.remove(old_path)
            original_name = secure_filename(f.filename)
            stored_name = f'{datetime.now().strftime("%Y%m%d%H%M%S%f")}_{original_name}'
            f.save(os.path.join(LEGAL_DIR, stored_name))
        now = datetime.now().isoformat(timespec='seconds')
        db.execute(
            '''UPDATE legal_nodes SET title=?, parent_id=?, description=?,
               stored_name=?, original_name=?, approval_status='pending',
               updated_at=? WHERE id=?''',
            (title, parent_id, description, stored_name, original_name, now, nid)
        )
        db.commit()
        flash('Đã cập nhật. Cần phê duyệt lại.', 'success')
        return redirect(url_for('legal_list'))

    return render_template('legal_form.html', options=options, action='edit',
                           node=dict(node), form=dict(node))


@app.route('/legal/<int:nid>/delete', methods=['POST'])
@login_required
def legal_delete(nid):
    db = get_db()
    node = db.execute('SELECT * FROM legal_nodes WHERE id=?', (nid,)).fetchone()
    if not node:
        abort(404)
    if node['stored_name']:
        fp = os.path.join(LEGAL_DIR, node['stored_name'])
        if os.path.exists(fp):
            os.remove(fp)
    db.execute('DELETE FROM legal_nodes WHERE id=?', (nid,))
    db.commit()
    flash('Đã xóa.', 'success')
    return redirect(url_for('legal_list'))


@app.route('/legal/<int:nid>/approve', methods=['POST'])
@login_required
@require_cap('approve_legal')
def legal_approve(nid):
    db = get_db()
    node = db.execute('SELECT * FROM legal_nodes WHERE id=?', (nid,)).fetchone()
    if not node:
        abort(404)
    now = datetime.now().isoformat(timespec='seconds')
    note = request.form.get('approval_note', '').strip()
    db.execute(
        '''UPDATE legal_nodes SET approval_status='approved', approved_by=?,
           approved_at=?, approval_note=?, updated_at=? WHERE id=?''',
        (current_user.id, now, note, now, nid)
    )
    creator_id = node['created_by']
    if creator_id != current_user.id:
        create_notification(creator_id,
                            f'Giấy tờ pháp lý «{node["title"]}» đã được PHÊ DUYỆT',
                            url_for('legal_list'))
    db.commit()
    flash('Đã phê duyệt giấy tờ pháp lý.', 'success')
    return redirect(url_for('legal_list'))


@app.route('/legal/<int:nid>/reject', methods=['POST'])
@login_required
@require_cap('approve_legal')
def legal_reject(nid):
    db = get_db()
    node = db.execute('SELECT * FROM legal_nodes WHERE id=?', (nid,)).fetchone()
    if not node:
        abort(404)
    now = datetime.now().isoformat(timespec='seconds')
    note = request.form.get('approval_note', '').strip()
    db.execute(
        '''UPDATE legal_nodes SET approval_status='rejected', approved_by=?,
           approved_at=?, approval_note=?, updated_at=? WHERE id=?''',
        (current_user.id, now, note, now, nid)
    )
    creator_id = node['created_by']
    if creator_id != current_user.id:
        create_notification(creator_id,
                            f'Giấy tờ pháp lý «{node["title"]}» bị TỪ CHỐI: {note}',
                            url_for('legal_list'))
    db.commit()
    flash('Đã từ chối.', 'warning')
    return redirect(url_for('legal_list'))


@app.route('/legal/<int:nid>/download')
@login_required
def legal_download(nid):
    db = get_db()
    node = db.execute('SELECT * FROM legal_nodes WHERE id=?', (nid,)).fetchone()
    if not node or not node['stored_name']:
        abort(404)
    return send_from_directory(LEGAL_DIR, node['stored_name'],
                               as_attachment=True,
                               download_name=node['original_name'] or node['stored_name'])


# ---------- Backup ----------
@app.route('/admin/backup')
@login_required
@admin_required
def admin_backup():
    mem = io.BytesIO()
    db_fd, db_tmp = tempfile.mkstemp(suffix='.db')
    os.close(db_fd)
    try:
        src = sqlite3.connect(DB_PATH)
        dst = sqlite3.connect(db_tmp)
        with dst:
            src.backup(dst)
        dst.close()
        src.close()
        with zipfile.ZipFile(mem, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(db_tmp, 'data.db')
            for root, _dirs, files in os.walk(UPLOAD_DIR):
                for fn in files:
                    full = os.path.join(root, fn)
                    zf.write(full, os.path.relpath(full, DATA_DIR))
    finally:
        try:
            os.remove(db_tmp)
        except OSError:
            pass
    mem.seek(0)
    name = 'backup_kbc_hp89_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.zip'
    return send_file(mem, as_attachment=True, download_name=name, mimetype='application/zip')


@app.route('/admin/db-status')
@login_required
@admin_required
def admin_db_status():
    disk_ok = os.path.exists('/var/data')
    db_exists = os.path.exists(DB_PATH)
    user_count = get_db().execute('SELECT COUNT(*) c FROM users').fetchone()['c']
    return f"""<pre style="font-family:monospace; padding:20px; font-size:14px">
=== DATABASE STATUS ===
DATA_DIR env var : {os.environ.get('DATA_DIR', '(not set)')}
DB_PATH          : {DB_PATH}
DB file exists   : {db_exists}
/var/data exists : {disk_ok}
Users in DB      : {user_count}

{'OK - database tren persistent disk' if '/var/data' in DB_PATH else 'CANH BAO: DB dang o thu muc tam, se mat khi redeploy!'}
</pre>"""


# ---------- CLI ----------
@app.cli.command('init-db')
def cli_init_db():
    init_db()
    print('DB initialized.')


# Khởi tạo DB khi import
init_db()


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=True)
